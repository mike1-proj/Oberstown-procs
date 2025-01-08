import os
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import NamedStyle
import openpyxl
from tkinter.messagebox import showinfo

"""Author Michael McNerney. This code version for Oberstown system is modified from the original to accommodate 
non-standard source file CSV formats found on Oberstown system.This code will read selected machine produced CSV files 
with non coded file names from a source folder and convert them into a Pandas data frame and then onto a user friendly 
Excel sheet data format with added identity file names and columns to make them recognisable and usable by others for 
later energy reports. 
I added code that uses openpyxl and named style function process to format columns of cells 
in the Excel file export into the correct date format. Openpyxl cannot format an entire set of column values,only 
individual cells. However, by clever use of a loop we can format all the cells in a chosen column and achieve 
the same result. Then, another piece of code is used at the end to get the column width correct for meter title column 
in each xl export file.This code version is for use with Oberstown BMS supervisor PC 
directory paths only. I have omitted the CSV export code so it only exports as Excel xlsx"""


def show_message():
    showinfo(message='The Process is completed! Please close the Data Log Processor Application before running again.')


# first we define the base path to the csv source folder which stays constant and create a var for each csv file name
# so,we can easily use the same base path and just add the file name we want at any stage.

base_path = r"C:\UnitronUC32\NCDFLUSK\Archive"

file_name1 = "D0080211.csv"
file_name2 = "D0061803.csv"
file_name3 = "D0061801.csv"
file_name4 = "D0061807.csv"
file_name5 = "D0061805.csv"
file_name6 = "D0051206.csv"
file_name7 = "D0051208.csv"
file_name8 = "D0051210.csv"
file_name9 = "D0050106.csv"
file_name10 = "D0050110.csv"
file_name11 = "D0041102.csv"
file_name12 = "D0041108.csv"
file_name13 = "D0041104.csv"
file_name14 = "D0041106.csv"
file_name15 = "D0031801.csv"
file_name16 = "D0031807.csv"
file_name17 = "D0031803.csv"
file_name18 = "D0031805.csv"
file_name19 = "D0020106.csv"
file_name20 = "D0020110.csv"
file_name21 = "D0021206.csv"
file_name22 = "D0021208.csv"
file_name23 = "D0021210.csv"
file_name24 = "D0011102.csv"
file_name25 = "D0011106.csv"
file_name26 = "D0011108.csv"
file_name27 = "D0011104.csv"
file_name28 = "D0071506.csv"
file_name29 = "D0071508.csv"
file_name30 = "D0071510.csv"
file_name31 = "D0070106.csv"
file_name32 = "D0071105.csv"
file_name33 = "D0081406.csv"
file_name34 = "D0081408.csv"
file_name35 = "D0081410.csv"
file_name36 = "D0080209.csv"
file_name37 = "D0110101.csv"
file_name38 = "D0110103.csv"
file_name39 = "D0110107.csv"
file_name40 = "D0110105.csv"
file_name41 = "D0120101.csv"
file_name42 = "D0120103.csv"
file_name43 = "D0090101.csv"
file_name44 = "D0090102.csv"
file_name45 = "D0090103.csv"
file_name46 = "D0090104.csv"
file_name47 = "D0090105.csv"
file_name48 = "D0090106.csv"
file_name49 = "D0090107.csv"
file_name50 = "D0090108.csv"
file_name51 = "D0090117.csv"
file_name52 = "D0100101.csv"
file_name53 = "D0100103.csv"
file_name54 = "D0100107.csv"

# set up the final destination folder for cleaned meter files.

base_dest_path = r"C:\Users\BMSuser\Desktop\Renamed_xlsx"
# set up the new column headers we will use in place of original file header for each new reformatted  file processed.
column_names = ['A', 'Intervals', '00:15', '00:30', '00:45', '01:00',
                '01:15', '01:30', '01:45', '02:00', '02:15', '02:30', '02:45', '03:00', '03:15', '03:30', '03:45',
                '04:00', '04:15', '04:30', '04:45', '05:00', '05:15', '05:30', '05:45', '06:00', '06:15', '06:30',
                '06:45', '07:00', '07:15', '07:30', '07:45', '08:00', '08:15', '08:30', '08:45', '09:00', '09:15',
                '09:30', '09:45', '10:00', '10:15', '10:30', '10:45', '11:00', '11:15', '11:30', '11:45', '12:00',
                '12:15', '12:30', '12:45', '13:00', '13:15', '13:30', '13:45', '14:00', '14:15', '14:30', '14:45',
                '15:00', '15:15', '15:30', '15:45', '16:00', '16:15', '16:30', '16:45', '17:00', '17:15', '17:30',
                '17:45', '18:00', '18:15', '18:30', '18:45', '19:00', '19:15', '19:30', '19:45', '20:00', '20:15',
                '20:30', '20:45', '21:00', '21:15', '21:30', '21:45', '22:00', '22:15', '22:30', '22:45', '23:00',
                '23:15', '23:30', '23:45', '24:00']

# now lets read the first file into a data frame from the data log target file ignoring the header row
# by use of "skip-rows" option because source file header data conflicts with body data columns in these files.
EdRec_gas_meter = pd.read_csv(base_path + "/" + file_name1, index_col=False, skiprows=[0])
# Now let's add our own custom column header names  to the data frame.
EdRec_gas_meter.columns = column_names
# now lets re-format the date column data from the source system csv which is formatted in a unique MS Excel date format
EdRec_gas_meter['A'] = pd.to_datetime(EdRec_gas_meter['A'], unit='D', origin='1899-12-30')
# now let's rename the "A" column to a more useful name when it is exported as an Excel file.
EdRec_gas_meter.rename(columns={"A": "EdRec_gas_mtr"}, inplace=True)
# now let's add a day total calculation column to the end of the columns
EdRec_gas_meter['Total_Day'] = EdRec_gas_meter['24:00'] - EdRec_gas_meter['00:15']

# Now the DF is as we want it. So, let's complete the formatting process for Excel export
# so that the DF exports as an Excel file and appears as we want it.

ewb = pd.ExcelWriter(base_dest_path + "/" + "EdRec_gas_meter.xlsx", engine="openpyxl")
nsmmyy = NamedStyle(name="cd1", number_format="DD-MM-YY")
nsmmmyy = NamedStyle(name="cd2", number_format="MMM-YY")
nsbyy = NamedStyle(name="cd3", number_format="MMMM-YY")
# now that the style has been set up for Excel writer we can  write the file.
EdRec_gas_meter.to_excel(excel_writer=ewb, sheet_name="EdRec_gas_meter")
# after this we have to format the EdRec_gas_meter column using the NamedStyles set up
ws = ewb.book["EdRec_gas_meter"]
for i in range(1, len(EdRec_gas_meter) + 2):
    ws.cell(row=i, column=2).style = nsmmyy
    # ws.cell(row=i, column=3).style = nsmmmyy
    # ws.cell(row=i, column=2).style = nsbyy
ewb.close()

# Finally we need to sort out the Unit 10 GS column width value, so it appears in the completed
# exported Excel sheet exactly as we want it with enough width for the title name.
worksheet = openpyxl.load_workbook(base_dest_path + "/" + "EdRec_gas_meter.xlsx")
sheet = worksheet.active
sheet.column_dimensions['B'].width = 19
worksheet.save(base_dest_path + "/" + "EdRec_gas_meter.xlsx")
worksheet.close()

# after this, we repeat the same process for every file to be processed in the old datalogs format. We simply change
# file name variable manually in the read path each time and skip the header row in the source csv file.

Unit_10_GS_meter = pd.read_csv(base_path + "/" + file_name2, index_col=False, skiprows=[0])
Unit_10_GS_meter.columns = column_names
Unit_10_GS_meter['A'] = pd.to_datetime(Unit_10_GS_meter['A'], unit='D', origin='1899-12-30')

Unit_10_GS_meter.rename(columns={"A": "Unit_10_GS"}, inplace=True)
Unit_10_GS_meter['Total_Day'] = Unit_10_GS_meter['24:00'] - Unit_10_GS_meter['00:15']

ewb = pd.ExcelWriter(base_dest_path + "/" + "Unit_10_GS_meter.xlsx", engine="openpyxl")
nsmmyy = NamedStyle(name="cd1", number_format="DD-MM-YY")
nsmmmyy = NamedStyle(name="cd2", number_format="MMM-YY")
nsbyy = NamedStyle(name="cd3", number_format="MMMM-YY")
# now that the style has been set up for Excel writer we can  write the file.
Unit_10_GS_meter.to_excel(excel_writer=ewb, sheet_name="Unit_10_GS")
# after this we have to format the Unit_10_GS_meter column using the NamedStyles set up
ws = ewb.book["Unit_10_GS"]
for i in range(1, len(Unit_10_GS_meter) + 2):
    ws.cell(row=i, column=2).style = nsmmyy
    # ws.cell(row=i, column=3).style = nsmmmyy
    # ws.cell(row=i, column=2).style = nsbyy
ewb.close()

# Finally we need to sort out the Unit 10 GS column width value in the exported Excel sheet
worksheet = openpyxl.load_workbook(base_dest_path + "/" + "Unit_10_GS_meter.xlsx")
sheet = worksheet.active
sheet.column_dimensions['B'].width = 19
worksheet.save(base_dest_path + "/" + "Unit_10_GS_meter.xlsx")
worksheet.close()

# next meter

Unit_10_lighting_meter = pd.read_csv(base_path + "/" + file_name3, index_col=False, skiprows=[0])
Unit_10_lighting_meter.columns = column_names
Unit_10_lighting_meter['A'] = pd.to_datetime(Unit_10_lighting_meter['A'], unit='D', origin='1899-12-30')

Unit_10_lighting_meter.rename(columns={"A": "Unit_10_light_meter"}, inplace=True)
Unit_10_lighting_meter['Total_Day'] = Unit_10_lighting_meter['24:00'] - Unit_10_lighting_meter['00:15']

# now let's write it as xlsx file format
ewb = pd.ExcelWriter(base_dest_path + "/" + "Unit_10_lighting_meter.xlsx", engine="openpyxl")
nsmmyy = NamedStyle(name="cd1", number_format="DD-MM-YY")
nsmmmyy = NamedStyle(name="cd2", number_format="MMM-YY")
nsbyy = NamedStyle(name="cd3", number_format="MMMM-YY")
# now that the style has been set up for Excel writer we can  write the file.
Unit_10_lighting_meter.to_excel(excel_writer=ewb, sheet_name="Unit_10_Light")
# after this we have to format the Unit_10_Light_meter column using the NamedStyles set up
ws = ewb.book["Unit_10_Light"]
for i in range(1, len(Unit_10_lighting_meter) + 2):
    ws.cell(row=i, column=2).style = nsmmyy
    # ws.cell(row=i, column=3).style = nsmmmyy
    # ws.cell(row=i, column=2).style = nsbyy
ewb.close()

# Finally we need to sort out the Unit_10_Light column width value in the completed Excel sheet
worksheet = openpyxl.load_workbook(base_dest_path + "/" + "Unit_10_lighting_meter.xlsx")
sheet = worksheet.active
sheet.column_dimensions['B'].width = 19
worksheet.save(base_dest_path + "/" + "Unit_10_lighting_meter.xlsx")
worksheet.close()

# next meter

Unit_10_Gas_meter = pd.read_csv(base_path + "/" + file_name4, index_col=False, skiprows=[0])
Unit_10_Gas_meter.columns = column_names
Unit_10_Gas_meter['A'] = pd.to_datetime(Unit_10_Gas_meter['A'], unit='D', origin='1899-12-30')

Unit_10_Gas_meter.rename(columns={"A": "Unit_10_Gas_Mtr"}, inplace=True)
Unit_10_Gas_meter['Total_Day'] = Unit_10_Gas_meter['24:00'] - Unit_10_Gas_meter['00:15']

# now let's write it as xlsx file format
ewb = pd.ExcelWriter(base_dest_path + "/" + "Unit_10_Gas_meter.xlsx", engine="openpyxl")
nsmmyy = NamedStyle(name="cd1", number_format="DD-MM-YY")
nsmmmyy = NamedStyle(name="cd2", number_format="MMM-YY")
nsbyy = NamedStyle(name="cd3", number_format="MMMM-YY")
# now that the style has been set up for Excel writer we can  write the file.
Unit_10_Gas_meter.to_excel(excel_writer=ewb, sheet_name="Unit_10_Gas")
# after this we have to format the Unit_10_Gas_meter column using the NamedStyles set up
ws = ewb.book["Unit_10_Gas"]
for i in range(1, len(Unit_10_Gas_meter) + 2):
    ws.cell(row=i, column=2).style = nsmmyy
    # ws.cell(row=i, column=3).style = nsmmmyy
    # ws.cell(row=i, column=2).style = nsbyy
ewb.close()

# Finally we need to sort out the Unit_10_Gas column width value in the completed Excel sheet
worksheet = openpyxl.load_workbook(base_dest_path + "/" + "Unit_10_Gas_meter.xlsx")
sheet = worksheet.active
sheet.column_dimensions['B'].width = 19
worksheet.save(base_dest_path + "/" + "Unit_10_Gas_meter.xlsx")
worksheet.close()

# next meter

Unit_10_SmallPower_meter = pd.read_csv(base_path + "/" + file_name5, index_col=False, skiprows=[0])
Unit_10_SmallPower_meter.columns = column_names
Unit_10_SmallPower_meter['A'] = pd.to_datetime(Unit_10_SmallPower_meter['A'], unit='D', origin='1899-12-30')

Unit_10_SmallPower_meter.rename(columns={"A": "Unit_10_Small_Pwr"}, inplace=True)
Unit_10_SmallPower_meter['Total_Day'] = Unit_10_SmallPower_meter['24:00'] - Unit_10_SmallPower_meter['00:15']

# now let's write it as xlsx file format
ewb = pd.ExcelWriter(base_dest_path + "/" + "Unit_10_SmallPower_meter.xlsx", engine="openpyxl")
nsmmyy = NamedStyle(name="cd1", number_format="DD-MM-YY")
nsmmmyy = NamedStyle(name="cd2", number_format="MMM-YY")
nsbyy = NamedStyle(name="cd3", number_format="MMMM-YY")
# now that the style has been set up for Excel writer we can  write the file.
Unit_10_SmallPower_meter.to_excel(excel_writer=ewb, sheet_name="Unit_10_SmallPwr")
# after this we have to format the Unit_10_SmallPower_meter column using the NamedStyles set up
ws = ewb.book["Unit_10_SmallPwr"]
for i in range(1, len(Unit_10_SmallPower_meter) + 2):
    ws.cell(row=i, column=2).style = nsmmyy
    # ws.cell(row=i, column=3).style = nsmmmyy
    # ws.cell(row=i, column=2).style = nsbyy
ewb.close()

# Finally we need to sort out the Unit_10_SmallPower_meter column width value in the exported Excel sheet
worksheet = openpyxl.load_workbook(base_dest_path + "/" + "Unit_10_SmallPower_meter.xlsx")
sheet = worksheet.active
sheet.column_dimensions['B'].width = 19
worksheet.save(base_dest_path + "/" + "Unit_10_SmallPower_meter.xlsx")
worksheet.close()

# next meter

Unit_9_Lighting_meter = pd.read_csv(base_path + "/" + file_name6, index_col=False, skiprows=[0])
Unit_9_Lighting_meter.columns = column_names
Unit_9_Lighting_meter['A'] = pd.to_datetime(Unit_9_Lighting_meter['A'], unit='D', origin='1899-12-30')

Unit_9_Lighting_meter.rename(columns={"A": "Unit_9_Lights"}, inplace=True)
Unit_9_Lighting_meter['Total_Day'] = Unit_9_Lighting_meter['24:00'] - Unit_9_Lighting_meter['00:15']

# now let's write it as xlsx file format
ewb = pd.ExcelWriter(base_dest_path + "/" + "Unit_9_Lighting_meter.xlsx", engine="openpyxl")
nsmmyy = NamedStyle(name="cd1", number_format="DD-MM-YY")
nsmmmyy = NamedStyle(name="cd2", number_format="MMM-YY")
nsbyy = NamedStyle(name="cd3", number_format="MMMM-YY")
# now that the style has been set up for Excel writer we can  write the file.
Unit_9_Lighting_meter.to_excel(excel_writer=ewb, sheet_name="Unit_9_Lighting")
# after this we have to format the Unit_9_Lighting_meter column using the NamedStyles set up
ws = ewb.book["Unit_9_Lighting"]
for i in range(1, len(Unit_9_Lighting_meter) + 2):
    ws.cell(row=i, column=2).style = nsmmyy
    # ws.cell(row=i, column=3).style = nsmmmyy
    # ws.cell(row=i, column=2).style = nsbyy
ewb.close()

# Finally we need to sort out the Unit_9_Lights column width value in the exported Excel sheet
worksheet = openpyxl.load_workbook(base_dest_path + "/" + "Unit_9_Lighting_meter.xlsx")
sheet = worksheet.active
sheet.column_dimensions['B'].width = 19
worksheet.save(base_dest_path + "/" + "Unit_9_Lighting_meter.xlsx")
worksheet.close()

# next meter

Unit_9_GS_meter = pd.read_csv(base_path + "/" + file_name7, index_col=False, skiprows=[0])
Unit_9_GS_meter.columns = column_names
Unit_9_GS_meter['A'] = pd.to_datetime(Unit_9_GS_meter['A'], unit='D', origin='1899-12-30')

Unit_9_GS_meter.rename(columns={"A": "Unit_9_GS_Mtr"}, inplace=True)
Unit_9_GS_meter['Total_Day'] = Unit_9_GS_meter['24:00'] - Unit_9_GS_meter['00:15']

# now let's write it as xlsx file format
ewb = pd.ExcelWriter(base_dest_path + "/" + "Unit_9_GS_meter.xlsx", engine="openpyxl")
nsmmyy = NamedStyle(name="cd1", number_format="DD-MM-YY")
nsmmmyy = NamedStyle(name="cd2", number_format="MMM-YY")
nsbyy = NamedStyle(name="cd3", number_format="MMMM-YY")
# now that the style has been set up for Excel writer we can  write the file.
Unit_9_GS_meter.to_excel(excel_writer=ewb, sheet_name="Unit_9_GS_Mtr")
# after this we have to format the Unit_9_GS_meter column using the NamedStyles set up
ws = ewb.book["Unit_9_GS_Mtr"]
for i in range(1, len(Unit_9_GS_meter) + 2):
    ws.cell(row=i, column=2).style = nsmmyy
    # ws.cell(row=i, column=3).style = nsmmmyy
    # ws.cell(row=i, column=2).style = nsbyy
ewb.close()

# Finally we need to sort out the Unit_9_GS column width value in the completed Excel sheet
worksheet = openpyxl.load_workbook(base_dest_path + "/" + "Unit_9_GS_meter.xlsx")
sheet = worksheet.active
sheet.column_dimensions['B'].width = 19
worksheet.save(base_dest_path + "/" + "Unit_9_GS_meter.xlsx")
worksheet.close()

# next meter

Unit_9_SmallPower_meter = pd.read_csv(base_path + "/" + file_name8, index_col=False, skiprows=[0])
Unit_9_SmallPower_meter.columns = column_names
Unit_9_SmallPower_meter['A'] = pd.to_datetime(Unit_9_SmallPower_meter['A'], unit='D', origin='1899-12-30')

Unit_9_SmallPower_meter.rename(columns={"A": "Unit_9_SmlPwr_Mtr"}, inplace=True)
Unit_9_SmallPower_meter['Total_Day'] = Unit_9_SmallPower_meter['24:00'] - Unit_9_SmallPower_meter['00:15']

# now let's write it as xlsx file format
ewb = pd.ExcelWriter(base_dest_path + "/" + "Unit_9_SmallPower_meter.xlsx", engine="openpyxl")
nsmmyy = NamedStyle(name="cd1", number_format="DD-MM-YY")
nsmmmyy = NamedStyle(name="cd2", number_format="MMM-YY")
nsbyy = NamedStyle(name="cd3", number_format="MMMM-YY")
# now that the style has been set up for Excel writer we can  write the file.
Unit_9_SmallPower_meter.to_excel(excel_writer=ewb, sheet_name="Unit_9_SmlPwr")
# after this we have to format the Unit_9_SMLPwr_meter column using the NamedStyles set up
ws = ewb.book["Unit_9_SmlPwr"]
for i in range(1, len(Unit_9_SmallPower_meter) + 2):
    ws.cell(row=i, column=2).style = nsmmyy
    # ws.cell(row=i, column=3).style = nsmmmyy
    # ws.cell(row=i, column=2).style = nsbyy
ewb.close()

# Finally we need to sort out the Unit_9_SmlPwr column width value in the completed Excel sheet
worksheet = openpyxl.load_workbook(base_dest_path + "/" + "Unit_9_SmallPower_meter.xlsx")
sheet = worksheet.active
sheet.column_dimensions['B'].width = 19
worksheet.save(base_dest_path + "/" + "Unit_9_SmallPower_meter.xlsx")
worksheet.close()

# next meter

Unit_9_Gas_meter = pd.read_csv(base_path + "/" + file_name9, index_col=False, skiprows=[0])
Unit_9_Gas_meter.columns = column_names
Unit_9_Gas_meter['A'] = pd.to_datetime(Unit_9_Gas_meter['A'], unit='D', origin='1899-12-30')

Unit_9_Gas_meter.rename(columns={"A": "Unit_9_Gas"}, inplace=True)
Unit_9_Gas_meter['Total_Day'] = Unit_9_Gas_meter['24:00'] - Unit_9_Gas_meter['00:15']

# now let's write it as xlsx file format
ewb = pd.ExcelWriter(base_dest_path + "/" + "Unit_9_Gas_meter.xlsx", engine="openpyxl")
nsmmyy = NamedStyle(name="cd1", number_format="DD-MM-YY")
nsmmmyy = NamedStyle(name="cd2", number_format="MMM-YY")
nsbyy = NamedStyle(name="cd3", number_format="MMMM-YY")
# now that the style has been set up for Excel writer we can  write the file.
Unit_9_Gas_meter.to_excel(excel_writer=ewb, sheet_name="Unit_9_Gas")
# after this we have to format the Unit_9_SMLPwr_meter column using the NamedStyles set up
ws = ewb.book["Unit_9_Gas"]
for i in range(1, len(Unit_9_Gas_meter) + 2):
    ws.cell(row=i, column=2).style = nsmmyy
    # ws.cell(row=i, column=3).style = nsmmmyy
    # ws.cell(row=i, column=2).style = nsbyy
ewb.close()

# Finally we need to sort out the Unit_9_SmlPwr column width value in the completed Excel sheet
worksheet = openpyxl.load_workbook(base_dest_path + "/" + "Unit_9_Gas_meter.xlsx")
sheet = worksheet.active
sheet.column_dimensions['B'].width = 19
worksheet.save(base_dest_path + "/" + "Unit_9_Gas_meter.xlsx")
worksheet.close()

# next meter

Unit_9_Water_meter = pd.read_csv(base_path + "/" + file_name10, index_col=False, skiprows=[0])
Unit_9_Water_meter.columns = column_names
Unit_9_Water_meter['A'] = pd.to_datetime(Unit_9_Water_meter['A'], unit='D', origin='1899-12-30')

Unit_9_Water_meter.rename(columns={"A": "Unit_9_Wtr_Mtr"}, inplace=True)
Unit_9_Water_meter['Total_Day'] = Unit_9_Water_meter['24:00'] - Unit_9_Water_meter['00:15']

# now let's write it as xlsx file format
ewb = pd.ExcelWriter(base_dest_path + "/" + "Unit_9_Water_meter.xlsx", engine="openpyxl")
nsmmyy = NamedStyle(name="cd1", number_format="DD-MM-YY")
nsmmmyy = NamedStyle(name="cd2", number_format="MMM-YY")
nsbyy = NamedStyle(name="cd3", number_format="MMMM-YY")
# now that the style has been set up for Excel writer we can  write the file.
Unit_9_Water_meter.to_excel(excel_writer=ewb, sheet_name="Unit_9_Wtr")
# after this we have to format the Unit_9_SMLPwr_meter column using the NamedStyles set up
ws = ewb.book["Unit_9_Wtr"]
for i in range(1, len(Unit_9_Water_meter) + 2):
    ws.cell(row=i, column=2).style = nsmmyy
    # ws.cell(row=i, column=3).style = nsmmmyy
    # ws.cell(row=i, column=2).style = nsbyy
ewb.close()

# Finally we need to sort out the Unit_9_SmlPwr column width value in the completed Excel sheet
worksheet = openpyxl.load_workbook(base_dest_path + "/" + "Unit_9_Water_meter.xlsx")
sheet = worksheet.active
sheet.column_dimensions['B'].width = 19
worksheet.save(base_dest_path + "/" + "Unit_9_Water_meter.xlsx")
worksheet.close()

# next meter

Unit_8_Lighting_meter = pd.read_csv(base_path + "/" + file_name11, index_col=False, skiprows=[0])
Unit_8_Lighting_meter.columns = column_names
Unit_8_Lighting_meter['A'] = pd.to_datetime(Unit_8_Lighting_meter['A'], unit='D', origin='1899-12-30')

Unit_8_Lighting_meter.rename(columns={"A": "Unit_8_Light_Mtr"}, inplace=True)
Unit_8_Lighting_meter['Total_Day'] = Unit_8_Lighting_meter['24:00'] - Unit_8_Lighting_meter['00:15']

# now let's write it as xlsx file format
ewb = pd.ExcelWriter(base_dest_path + "/" + "Unit_8_Lighting_meter.xlsx", engine="openpyxl")
nsmmyy = NamedStyle(name="cd1", number_format="DD-MM-YY")
nsmmmyy = NamedStyle(name="cd2", number_format="MMM-YY")
nsbyy = NamedStyle(name="cd3", number_format="MMMM-YY")
# now that the style has been set up for Excel writer we can  write the file.
Unit_8_Lighting_meter.to_excel(excel_writer=ewb, sheet_name="Unit_8_Light")
# after this we have to format the Unit_9_SMLPwr_meter column using the NamedStyles set up
ws = ewb.book["Unit_8_Light"]
for i in range(1, len(Unit_8_Lighting_meter) + 2):
    ws.cell(row=i, column=2).style = nsmmyy
    # ws.cell(row=i, column=3).style = nsmmmyy
    # ws.cell(row=i, column=2).style = nsbyy
ewb.close()

# Finally we need to sort out the Unit_9_SmlPwr column width value in the completed Excel sheet
worksheet = openpyxl.load_workbook(base_dest_path + "/" + "Unit_8_Lighting_meter.xlsx")
sheet = worksheet.active
sheet.column_dimensions['B'].width = 19
worksheet.save(base_dest_path + "/" + "Unit_8_Lighting_meter.xlsx")
worksheet.close()

# next meter

Unit_8_Gas_meter = pd.read_csv(base_path + "/" + file_name12, index_col=False, skiprows=[0])
Unit_8_Gas_meter.columns = column_names
Unit_8_Gas_meter['A'] = pd.to_datetime(Unit_8_Gas_meter['A'], unit='D', origin='1899-12-30')

Unit_8_Gas_meter.rename(columns={"A": "Unit_8_Gas_Mtr"}, inplace=True)
Unit_8_Gas_meter['Total_Day'] = Unit_8_Gas_meter['24:00'] - Unit_8_Gas_meter['00:15']

# now let's write it as xlsx file format
ewb = pd.ExcelWriter(base_dest_path + "/" + "Unit_8_Gas_meter.xlsx", engine="openpyxl")
nsmmyy = NamedStyle(name="cd1", number_format="DD-MM-YY")
nsmmmyy = NamedStyle(name="cd2", number_format="MMM-YY")
nsbyy = NamedStyle(name="cd3", number_format="MMMM-YY")
# now that the style has been set up for Excel Writer we can  write the file.
Unit_8_Gas_meter.to_excel(excel_writer=ewb, sheet_name="Unit_8_Gas_Mtr")
# after this we have to format the Unit_9_SMLPwr_meter column using the NamedStyles set up
ws = ewb.book["Unit_8_Gas_Mtr"]
for i in range(1, len(Unit_8_Gas_meter) + 2):
    ws.cell(row=i, column=2).style = nsmmyy
    # ws.cell(row=i, column=3).style = nsmmmyy
    # ws.cell(row=i, column=2).style = nsbyy
ewb.close()

# Finally we need to sort out the Unit_9_SmlPwr column width value in the completed Excel sheet
worksheet = openpyxl.load_workbook(base_dest_path + "/" + "Unit_8_Gas_meter.xlsx")
sheet = worksheet.active
sheet.column_dimensions['B'].width = 19
worksheet.save(base_dest_path + "/" + "Unit_8_Gas_meter.xlsx")
worksheet.close()

# next meter

Unit_8_Gs_meter = pd.read_csv(base_path + "/" + file_name13, index_col=False, skiprows=[0])
Unit_8_Gs_meter.columns = column_names
Unit_8_Gs_meter['A'] = pd.to_datetime(Unit_8_Gs_meter['A'], unit='D', origin='1899-12-30')

Unit_8_Gs_meter.rename(columns={"A": "Unit_8_GS_Mtr"}, inplace=True)
Unit_8_Gs_meter['Total_Day'] = Unit_8_Gs_meter['24:00'] - Unit_8_Gs_meter['00:15']

# now let's write it as xlsx file format
ewb = pd.ExcelWriter(base_dest_path + "/" + "Unit_8_Gs_meter.xlsx", engine="openpyxl")
nsmmyy = NamedStyle(name="cd1", number_format="DD-MM-YY")
nsmmmyy = NamedStyle(name="cd2", number_format="MMM-YY")
nsbyy = NamedStyle(name="cd3", number_format="MMMM-YY")
# now that the style has been set up for Excel Writer we can  write the file.
Unit_8_Gs_meter.to_excel(excel_writer=ewb, sheet_name="Unit_8_Gs_Mtr")
# after this we have to format the Unit_8_Gs_meter column using the NamedStyles set up
ws = ewb.book["Unit_8_Gs_Mtr"]
for i in range(1, len(Unit_8_Gs_meter) + 2):
    ws.cell(row=i, column=2).style = nsmmyy
    # ws.cell(row=i, column=3).style = nsmmmyy
    # ws.cell(row=i, column=2).style = nsbyy
ewb.close()

# Finally we need to sort out the Unit_8_Gs_meter column width value in the completed Excel sheet
worksheet = openpyxl.load_workbook(base_dest_path + "/" + "Unit_8_Gs_meter.xlsx")
sheet = worksheet.active
sheet.column_dimensions['B'].width = 19
worksheet.save(base_dest_path + "/" + "Unit_8_Gs_meter.xlsx")
worksheet.close()

# next meter

Unit_8_SmallPower_meter = pd.read_csv(base_path + "/" + file_name14, index_col=False, skiprows=[0])
Unit_8_SmallPower_meter.columns = column_names
Unit_8_SmallPower_meter['A'] = pd.to_datetime(Unit_8_SmallPower_meter['A'], unit='D', origin='1899-12-30')

Unit_8_SmallPower_meter.rename(columns={"A": "Unit_8_SmallPwr_Mtr"}, inplace=True)
Unit_8_SmallPower_meter['Total_Day'] = Unit_8_SmallPower_meter['24:00'] - Unit_8_SmallPower_meter['00:15']

# now let's write it as xlsx file format
ewb = pd.ExcelWriter(base_dest_path + "/" + "Unit_8_SmallPower_meter.xlsx", engine="openpyxl")
nsmmyy = NamedStyle(name="cd1", number_format="DD-MM-YY")
nsmmmyy = NamedStyle(name="cd2", number_format="MMM-YY")
nsbyy = NamedStyle(name="cd3", number_format="MMMM-YY")
# now that the style has been set up for Excel Writer we can  write the file.
Unit_8_SmallPower_meter.to_excel(excel_writer=ewb, sheet_name="Unit_8_SmlPwr")
# after this we have to format the Unit_9_SMLPwr_meter column using the NamedStyles set up
ws = ewb.book["Unit_8_SmlPwr"]
for i in range(1, len(Unit_8_SmallPower_meter) + 2):
    ws.cell(row=i, column=2).style = nsmmyy
    # ws.cell(row=i, column=3).style = nsmmmyy
    # ws.cell(row=i, column=2).style = nsbyy
ewb.close()

# Finally we need to sort out the Unit_9_SmlPwr column width value in the completed Excel sheet
worksheet = openpyxl.load_workbook(base_dest_path + "/" + "Unit_8_SmallPower_meter.xlsx")
sheet = worksheet.active
sheet.column_dimensions['B'].width = 19
worksheet.save(base_dest_path + "/" + "Unit_8_SmallPower_meter.xlsx")
worksheet.close()

# next meter

Unit_7_Lighting_meter = pd.read_csv(base_path + "/" + file_name15, index_col=False, skiprows=[0])
Unit_7_Lighting_meter.columns = column_names
Unit_7_Lighting_meter['A'] = pd.to_datetime(Unit_7_Lighting_meter['A'], unit='D', origin='1899-12-30')

Unit_7_Lighting_meter.rename(columns={"A": "Unit_7_Light"}, inplace=True)
Unit_7_Lighting_meter['Total_Day'] = Unit_7_Lighting_meter['24:00'] - Unit_7_Lighting_meter['00:15']

# now let's write it as xlsx file format
ewb = pd.ExcelWriter(base_dest_path + "/" + "Unit_7_Lighting_meter.xlsx", engine="openpyxl")
nsmmyy = NamedStyle(name="cd1", number_format="DD-MM-YY")
nsmmmyy = NamedStyle(name="cd2", number_format="MMM-YY")
nsbyy = NamedStyle(name="cd3", number_format="MMMM-YY")
# now that the style has been set up for Excel Writer we can  write the file.
Unit_7_Lighting_meter.to_excel(excel_writer=ewb, sheet_name="Unit_7_Light")
# after this we have to format the Unit_7_Lighting_meter column using the NamedStyles set up
ws = ewb.book["Unit_7_Light"]
for i in range(1, len(Unit_7_Lighting_meter) + 2):
    ws.cell(row=i, column=2).style = nsmmyy
    # ws.cell(row=i, column=3).style = nsmmmyy
    # ws.cell(row=i, column=2).style = nsbyy
ewb.close()

# Finally we need to sort out the Unit_9_SmlPwr column width value in the completed Excel sheet
worksheet = openpyxl.load_workbook(base_dest_path + "/" + "Unit_7_Lighting_meter.xlsx")
sheet = worksheet.active
sheet.column_dimensions['B'].width = 19
worksheet.save(base_dest_path + "/" + "Unit_7_Lighting_meter.xlsx")
worksheet.close()

# next meter

Unit_7_Gas_meter = pd.read_csv(base_path + "/" + file_name16, index_col=False, skiprows=[0])
Unit_7_Gas_meter.columns = column_names
Unit_7_Gas_meter['A'] = pd.to_datetime(Unit_7_Gas_meter['A'], unit='D', origin='1899-12-30')

Unit_7_Gas_meter.rename(columns={"A": "Unit_7_Gas_Mtr"}, inplace=True)
Unit_7_Gas_meter['Total_Day'] = Unit_7_Gas_meter['24:00'] - Unit_7_Gas_meter['00:15']

# now let's write it as xlsx file format
ewb = pd.ExcelWriter(base_dest_path + "/" + "Unit_7_Gas_meter.xlsx", engine="openpyxl")
nsmmyy = NamedStyle(name="cd1", number_format="DD-MM-YY")
nsmmmyy = NamedStyle(name="cd2", number_format="MMM-YY")
nsbyy = NamedStyle(name="cd3", number_format="MMMM-YY")
# now that the style has been set up for Excel Writer we can  write the file.
Unit_7_Gas_meter.to_excel(excel_writer=ewb, sheet_name="Unit_7_Gas_Mtr")
# after this we have to format the Unit_7_Lighting_meter column using the NamedStyles set up
ws = ewb.book["Unit_7_Gas_Mtr"]
for i in range(1, len(Unit_7_Gas_meter) + 2):
    ws.cell(row=i, column=2).style = nsmmyy
    # ws.cell(row=i, column=3).style = nsmmmyy
    # ws.cell(row=i, column=2).style = nsbyy
ewb.close()

# Finally we need to sort out the Unit_9_SmlPwr column width value in the completed Excel sheet
worksheet = openpyxl.load_workbook(base_dest_path + "/" + "Unit_7_Gas_meter.xlsx")
sheet = worksheet.active
sheet.column_dimensions['B'].width = 19
worksheet.save(base_dest_path + "/" + "Unit_7_Gas_meter.xlsx")
worksheet.close()

# next meter

Unit_7_GS_meter = pd.read_csv(base_path + "/" + file_name17, index_col=False, skiprows=[0])
Unit_7_GS_meter.columns = column_names
Unit_7_GS_meter['A'] = pd.to_datetime(Unit_7_GS_meter['A'], unit='D', origin='1899-12-30')

Unit_7_GS_meter.rename(columns={"A": "Unit_7_GS"}, inplace=True)
Unit_7_GS_meter['Total_Day'] = Unit_7_GS_meter['24:00'] - Unit_7_GS_meter['00:15']

# now let's write it as xlsx file format
ewb = pd.ExcelWriter(base_dest_path + "/" + "Unit_7_Gs_meter.xlsx", engine="openpyxl")
nsmmyy = NamedStyle(name="cd1", number_format="DD-MM-YY")
nsmmmyy = NamedStyle(name="cd2", number_format="MMM-YY")
nsbyy = NamedStyle(name="cd3", number_format="MMMM-YY")
# now that the style has been set up for Excel Writer we can  write the file.
Unit_7_GS_meter.to_excel(excel_writer=ewb, sheet_name="Unit_7_Gs")
# after this we have to format the Unit_7_Lighting_meter column using the NamedStyles set up
ws = ewb.book["Unit_7_Gs"]
for i in range(1, len(Unit_7_GS_meter) + 2):
    ws.cell(row=i, column=2).style = nsmmyy
    # ws.cell(row=i, column=3).style = nsmmmyy
    # ws.cell(row=i, column=2).style = nsbyy
ewb.close()

# Finally we need to sort out the Unit_9_SmlPwr column width value in the completed Excel sheet
worksheet = openpyxl.load_workbook(base_dest_path + "/" + "Unit_7_Gs_meter.xlsx")
sheet = worksheet.active
sheet.column_dimensions['B'].width = 19
worksheet.save(base_dest_path + "/" + "Unit_7_Gs_meter.xlsx")
worksheet.close()

# next meter

Unit_7_SmallPower_meter = pd.read_csv(base_path + "/" + file_name18, index_col=False, skiprows=[0])
Unit_7_SmallPower_meter.columns = column_names
Unit_7_SmallPower_meter['A'] = pd.to_datetime(Unit_7_SmallPower_meter['A'], unit='D', origin='1899-12-30')

Unit_7_SmallPower_meter.rename(columns={"A": "Unit_7_SmlPwr_Mtr"}, inplace=True)
Unit_7_SmallPower_meter['Total_Day'] = Unit_7_SmallPower_meter['24:00'] - Unit_7_SmallPower_meter['00:15']

# now let's write it as xlsx file format
ewb = pd.ExcelWriter(base_dest_path + "/" + "Unit_7_SmallPower_meter.xlsx", engine="openpyxl")
nsmmyy = NamedStyle(name="cd1", number_format="DD-MM-YY")
nsmmmyy = NamedStyle(name="cd2", number_format="MMM-YY")
nsbyy = NamedStyle(name="cd3", number_format="MMMM-YY")
# now that the style has been set up for Excel Writer we can  write the file.
Unit_7_SmallPower_meter.to_excel(excel_writer=ewb, sheet_name="Unit_7_SmlPwr")
# after this we have to format the Unit_7_Lighting_meter column using the NamedStyles set up
ws = ewb.book["Unit_7_SmlPwr"]
for i in range(1, len(Unit_7_SmallPower_meter) + 2):
    ws.cell(row=i, column=2).style = nsmmyy
    # ws.cell(row=i, column=3).style = nsmmmyy
    # ws.cell(row=i, column=2).style = nsbyy
ewb.close()

# Finally we need to sort out the Unit_9_SmlPwr column width value in the completed Excel sheet
worksheet = openpyxl.load_workbook(base_dest_path + "/" + "Unit_7_SmallPower_meter.xlsx")
sheet = worksheet.active
sheet.column_dimensions['B'].width = 19
worksheet.save(base_dest_path + "/" + "Unit_7_SmallPower_meter.xlsx")
worksheet.close()

# next meter

Unit_6_Gas_meter = pd.read_csv(base_path + "/" + file_name19, index_col=False, skiprows=[0])
Unit_6_Gas_meter.columns = column_names
Unit_6_Gas_meter['A'] = pd.to_datetime(Unit_6_Gas_meter['A'], unit='D', origin='1899-12-30')

Unit_6_Gas_meter.rename(columns={"A": "Unit_6_Gas_Mtr"}, inplace=True)
Unit_6_Gas_meter['Total_Day'] = Unit_6_Gas_meter['24:00'] - Unit_6_Gas_meter['00:15']

# now let's write it as xlsx file format
ewb = pd.ExcelWriter(base_dest_path + "/" + "Unit_6_Gas_meter.xlsx", engine="openpyxl")
nsmmyy = NamedStyle(name="cd1", number_format="DD-MM-YY")
nsmmmyy = NamedStyle(name="cd2", number_format="MMM-YY")
nsbyy = NamedStyle(name="cd3", number_format="MMMM-YY")
# now that the style has been set up for Excel Writer we can  write the file.
Unit_6_Gas_meter.to_excel(excel_writer=ewb, sheet_name="Unit_6_Gas")
# after this we have to format the Unit_6_Gas_meter column using the NamedStyles set up
ws = ewb.book["Unit_6_Gas"]
for i in range(1, len(Unit_6_Gas_meter) + 2):
    ws.cell(row=i, column=2).style = nsmmyy
    # ws.cell(row=i, column=3).style = nsmmmyy
    # ws.cell(row=i, column=2).style = nsbyy
ewb.close()

# Finally we need to sort out the Unit_9_SmlPwr column width value in the completed Excel sheet
worksheet = openpyxl.load_workbook(base_dest_path + "/" + "Unit_6_Gas_meter.xlsx")
sheet = worksheet.active
sheet.column_dimensions['B'].width = 19
worksheet.save(base_dest_path + "/" + "Unit_6_Gas_meter.xlsx")
worksheet.close()

# next meter

Unit_6_Water_meter = pd.read_csv(base_path + "/" + file_name20, index_col=False, skiprows=[0])
Unit_6_Water_meter.columns = column_names
Unit_6_Water_meter['A'] = pd.to_datetime(Unit_6_Water_meter['A'], unit='D', origin='1899-12-30')

Unit_6_Water_meter.rename(columns={"A": "Unit_6_Wtr_Mtr"}, inplace=True)
Unit_6_Water_meter['Total_Day'] = Unit_6_Water_meter['24:00'] - Unit_6_Water_meter['00:15']

# now let's write it as xlsx file format
ewb = pd.ExcelWriter(base_dest_path + "/" + "Unit_6_Water_meter.xlsx", engine="openpyxl")
nsmmyy = NamedStyle(name="cd1", number_format="DD-MM-YY")
nsmmmyy = NamedStyle(name="cd2", number_format="MMM-YY")
nsbyy = NamedStyle(name="cd3", number_format="MMMM-YY")
# now that the style has been set up for Excel Writer we can  write the file.
Unit_6_Water_meter.to_excel(excel_writer=ewb, sheet_name="Unit_6_Water")
# after this we have to format the Unit_6_Water_meter column using the NamedStyles set up
ws = ewb.book["Unit_6_Water"]
for i in range(1, len(Unit_6_Water_meter) + 2):
    ws.cell(row=i, column=2).style = nsmmyy
    # ws.cell(row=i, column=3).style = nsmmmyy
    # ws.cell(row=i, column=2).style = nsbyy
ewb.close()

# Finally we need to sort out the Unit_6_Water_Meter column width value in the completed Excel sheet
worksheet = openpyxl.load_workbook(base_dest_path + "/" + "Unit_6_Water_meter.xlsx")
sheet = worksheet.active
sheet.column_dimensions['B'].width = 19
worksheet.save(base_dest_path + "/" + "Unit_6_Water_meter.xlsx")
worksheet.close()

# next meter

Unit_6_Lighting_meter = pd.read_csv(base_path + "/" + file_name21, index_col=False, skiprows=[0])
Unit_6_Lighting_meter.columns = column_names
Unit_6_Lighting_meter['A'] = pd.to_datetime(Unit_6_Lighting_meter['A'], unit='D', origin='1899-12-30')

Unit_6_Lighting_meter.rename(columns={"A": "Unit_6_Light"}, inplace=True)
Unit_6_Lighting_meter['Total_Day'] = Unit_6_Lighting_meter['24:00'] - Unit_6_Lighting_meter['00:15']

# now let's write it as xlsx file format
ewb = pd.ExcelWriter(base_dest_path + "/" + "Unit_6_Lighting_meter.xlsx", engine="openpyxl")
nsmmyy = NamedStyle(name="cd1", number_format="DD-MM-YY")
nsmmmyy = NamedStyle(name="cd2", number_format="MMM-YY")
nsbyy = NamedStyle(name="cd3", number_format="MMMM-YY")
# now that the style has been set up for Excel Writer we can  write the file.
Unit_6_Lighting_meter.to_excel(excel_writer=ewb, sheet_name="Unit_6_Light")
# after this we have to format the Unit_6_Gas_meter column using the NamedStyles set up
ws = ewb.book["Unit_6_Light"]
for i in range(1, len(Unit_6_Lighting_meter) + 2):
    ws.cell(row=i, column=2).style = nsmmyy
    # ws.cell(row=i, column=3).style = nsmmmyy
    # ws.cell(row=i, column=2).style = nsbyy
ewb.close()

# Finally we need to sort out the Unit_9_SmlPwr column width value in the completed Excel sheet
worksheet = openpyxl.load_workbook(base_dest_path + "/" + "Unit_6_Lighting_meter.xlsx")
sheet = worksheet.active
sheet.column_dimensions['B'].width = 19
worksheet.save(base_dest_path + "/" + "Unit_6_Lighting_meter.xlsx")
worksheet.close()

# next meter

Unit_6_GS_meter = pd.read_csv(base_path + "/" + file_name22, index_col=False, skiprows=[0])
Unit_6_GS_meter.columns = column_names
Unit_6_GS_meter['A'] = pd.to_datetime(Unit_6_GS_meter['A'], unit='D', origin='1899-12-30')

Unit_6_GS_meter.rename(columns={"A": "Unit_6_GS_Mtr"}, inplace=True)
Unit_6_GS_meter['Total_Day'] = Unit_6_GS_meter['24:00'] - Unit_6_GS_meter['00:15']

# now let's write it as xlsx file format
# first we need to set up some format styles to be applied to the Excel sheet date column using the named_style module
ewb = pd.ExcelWriter(base_dest_path + "/" + "Unit_6_GS_meter.xlsx", engine="openpyxl")
nsmmyy = NamedStyle(name="cd1", number_format="DD-MM-YY")
nsmmmyy = NamedStyle(name="cd2", number_format="MMM-YY")
nsbyy = NamedStyle(name="cd3", number_format="MMMM-YY")
# now that the style has been set up for Excel Writer we can  write the file.
Unit_6_GS_meter.to_excel(excel_writer=ewb, sheet_name="Unit_6_GS_Mtr")
# after this we have to format the Unit_6_GS_meter date column using the NamedStyles set up
ws = ewb.book["Unit_6_GS_Mtr"]
for i in range(1, len(Unit_6_GS_meter) + 2):
    ws.cell(row=i, column=2).style = nsmmyy
    # ws.cell(row=i, column=3).style = nsmmmyy
    # ws.cell(row=i, column=2).style = nsbyy
ewb.close()

# Finally we need to sort out the Unit_6_GS_meter column width value in the completed Excel sheet
worksheet = openpyxl.load_workbook(base_dest_path + "/" + "Unit_6_GS_meter.xlsx")
sheet = worksheet.active
sheet.column_dimensions['B'].width = 19
worksheet.save(base_dest_path + "/" + "Unit_6_GS_meter.xlsx")
worksheet.close()

# next meter

Unit_6_SmallPower_meter = pd.read_csv(base_path + "/" + file_name23, index_col=False, skiprows=[0])
Unit_6_SmallPower_meter.columns = column_names
Unit_6_SmallPower_meter['A'] = pd.to_datetime(Unit_6_SmallPower_meter['A'], unit='D', origin='1899-12-30')

Unit_6_SmallPower_meter.rename(columns={"A": "Unit_6_SmlPwr"}, inplace=True)
Unit_6_SmallPower_meter['Total_Day'] = Unit_6_SmallPower_meter['24:00'] - Unit_6_SmallPower_meter['00:15']

# now let's write it as xlsx file format
# first we need to set up some format styles to be applied to the Excel sheet date column using the named_style module
ewb = pd.ExcelWriter(base_dest_path + "/" + "Unit_6_SmallPower_meter.xlsx", engine="openpyxl")
nsmmyy = NamedStyle(name="cd1", number_format="DD-MM-YY")
nsmmmyy = NamedStyle(name="cd2", number_format="MMM-YY")
nsbyy = NamedStyle(name="cd3", number_format="MMMM-YY")
# now that the style has been set up for Excel Writer we can  write the file.
Unit_6_SmallPower_meter.to_excel(excel_writer=ewb, sheet_name="Unit_6_SmlPwr")
# after this we have to format the Unit_6_SmallPower_meter date column using the NamedStyles set up
ws = ewb.book["Unit_6_SmlPwr"]
for i in range(1, len(Unit_6_SmallPower_meter) + 2):
    ws.cell(row=i, column=2).style = nsmmyy
    # ws.cell(row=i, column=3).style = nsmmmyy
    # ws.cell(row=i, column=2).style = nsbyy
ewb.close()

# Finally we need to sort out the Unit_6_SmallPower_meter column width value in the completed Excel sheet
worksheet = openpyxl.load_workbook(base_dest_path + "/" + "Unit_6_SmallPower_meter.xlsx")
sheet = worksheet.active
sheet.column_dimensions['B'].width = 19
worksheet.save(base_dest_path + "/" + "Unit_6_SmallPower_meter.xlsx")
worksheet.close()

# next meter

Unit_5_Gas_meter = pd.read_csv(base_path + "/" + file_name24, index_col=False, skiprows=[0])
Unit_5_Gas_meter.columns = column_names
Unit_5_Gas_meter['A'] = pd.to_datetime(Unit_5_Gas_meter['A'], unit='D', origin='1899-12-30')

Unit_5_Gas_meter.rename(columns={"A": "Unit_10_GS"}, inplace=True)
Unit_5_Gas_meter['Total_Day'] = Unit_5_Gas_meter['24:00'] - Unit_5_Gas_meter['00:15']

# now let's write it as xlsx file format
# first we need to set up some format styles to be applied to the Excel sheet date column using the named_style module
ewb = pd.ExcelWriter(base_dest_path + "/" + "Unit_5_Gas_meter.xlsx", engine="openpyxl")
nsmmyy = NamedStyle(name="cd1", number_format="DD-MM-YY")
nsmmmyy = NamedStyle(name="cd2", number_format="MMM-YY")
nsbyy = NamedStyle(name="cd3", number_format="MMMM-YY")
# now that the style has been set up for Excel Writer we can  write the file.
Unit_5_Gas_meter.to_excel(excel_writer=ewb, sheet_name="Unit_5_Gas")
# after this we have to format the Unit_5_Gas_meter date column using the NamedStyles set up
ws = ewb.book["Unit_5_Gas"]
for i in range(1, len(Unit_5_Gas_meter) + 2):
    ws.cell(row=i, column=2).style = nsmmyy
    # ws.cell(row=i, column=3).style = nsmmmyy
    # ws.cell(row=i, column=2).style = nsbyy
ewb.close()

# Finally we need to sort out the Unit_6_Gas_meter column width value in the completed Excel sheet
worksheet = openpyxl.load_workbook(base_dest_path + "/" + "Unit_5_Gas_meter.xlsx")
sheet = worksheet.active
sheet.column_dimensions['B'].width = 19
worksheet.save(base_dest_path + "/" + "Unit_5_Gas_meter.xlsx")
worksheet.close()

# next meter

Unit_5_Gs_meter = pd.read_csv(base_path + "/" + file_name25, index_col=False, skiprows=[0])
Unit_5_Gs_meter.columns = column_names
Unit_5_Gs_meter['A'] = pd.to_datetime(Unit_5_Gs_meter['A'], unit='D', origin='1899-12-30')

Unit_5_Gs_meter.rename(columns={"A": "Unit_5_GS_Mtr"}, inplace=True)
Unit_5_Gs_meter['Total_Day'] = Unit_5_Gs_meter['24:00'] - Unit_5_Gs_meter['00:15']

# now let's write it as xlsx file format
# first we need to set up some format styles to be applied to the Excel sheet date column using the named_style module
ewb = pd.ExcelWriter(base_dest_path + "/" + "Unit_5_Gs_meter.xlsx", engine="openpyxl")
nsmmyy = NamedStyle(name="cd1", number_format="DD-MM-YY")
nsmmmyy = NamedStyle(name="cd2", number_format="MMM-YY")
nsbyy = NamedStyle(name="cd3", number_format="MMMM-YY")
# now that the style has been set up for Excel Writer we can  write the file.
Unit_5_Gs_meter.to_excel(excel_writer=ewb, sheet_name="Unit_5_Gs")
# after this we have to format the Unit_5_Gs_meter date column using the NamedStyles set up
ws = ewb.book["Unit_5_Gs"]
for i in range(1, len(Unit_5_Gas_meter) + 2):
    ws.cell(row=i, column=2).style = nsmmyy
    # ws.cell(row=i, column=3).style = nsmmmyy
    # ws.cell(row=i, column=2).style = nsbyy
ewb.close()

# Finally we need to sort out the Unit_5_Gs_meter column width value in the completed Excel sheet
worksheet = openpyxl.load_workbook(base_dest_path + "/" + "Unit_5_Gs_meter.xlsx")
sheet = worksheet.active
sheet.column_dimensions['B'].width = 19
worksheet.save(base_dest_path + "/" + "Unit_5_Gs_meter.xlsx")
worksheet.close()

# next meter

Unit_5_SmallPower_meter = pd.read_csv(base_path + "/" + file_name26, index_col=False, skiprows=[0])
Unit_5_SmallPower_meter.columns = column_names
Unit_5_SmallPower_meter['A'] = pd.to_datetime(Unit_5_SmallPower_meter['A'], unit='D', origin='1899-12-30')

Unit_5_SmallPower_meter.rename(columns={"A": "Unit_5_SmlPwr"}, inplace=True)
Unit_5_SmallPower_meter['Total_Day'] = Unit_5_SmallPower_meter['24:00'] - Unit_5_SmallPower_meter['00:15']

# now let's write it as xlsx file format
# first we need to set up some format styles to be applied to the Excel sheet date column using the named_style module
ewb = pd.ExcelWriter(base_dest_path + "/" + "Unit_5_SmallPower_meter.xlsx", engine="openpyxl")
nsmmyy = NamedStyle(name="cd1", number_format="DD-MM-YY")
nsmmmyy = NamedStyle(name="cd2", number_format="MMM-YY")
nsbyy = NamedStyle(name="cd3", number_format="MMMM-YY")
# now that the style has been set up for Excel Writer we can  write the file.
Unit_5_SmallPower_meter.to_excel(excel_writer=ewb, sheet_name="Unit_5_SmlPwr")
# after this we have to format the Unit_5_SmallPower_meter date column using the NamedStyles set up
ws = ewb.book["Unit_5_SmlPwr"]
for i in range(1, len(Unit_5_SmallPower_meter) + 2):
    ws.cell(row=i, column=2).style = nsmmyy
    # ws.cell(row=i, column=3).style = nsmmmyy
    # ws.cell(row=i, column=2).style = nsbyy
ewb.close()

# Finally we need to sort out the Unit_5_SmallPower_meter column width value in the completed Excel sheet
worksheet = openpyxl.load_workbook(base_dest_path + "/" + "Unit_5_SmallPower_meter.xlsx")
sheet = worksheet.active
sheet.column_dimensions['B'].width = 19
worksheet.save(base_dest_path + "/" + "Unit_5_SmallPower_meter.xlsx")
worksheet.close()

# next meter

Unit_5_Lighting_meter = pd.read_csv(base_path + "/" + file_name27, index_col=False, skiprows=[0])
Unit_5_Lighting_meter.columns = column_names
Unit_5_Lighting_meter['A'] = pd.to_datetime(Unit_5_Lighting_meter['A'], unit='D', origin='1899-12-30')

Unit_5_Lighting_meter.rename(columns={"A": "Unit_5_Light"}, inplace=True)
Unit_5_Lighting_meter['Total_Day'] = Unit_5_Lighting_meter['24:00'] - Unit_5_Lighting_meter['00:15']

# now let's write it as xlsx file format
# first we need to set up some format styles to be applied to the Excel sheet date column using the named_style module
ewb = pd.ExcelWriter(base_dest_path + "/" + "Unit_5_Lighting_meter.xlsx", engine="openpyxl")
nsmmyy = NamedStyle(name="cd1", number_format="DD-MM-YY")
nsmmmyy = NamedStyle(name="cd2", number_format="MMM-YY")
nsbyy = NamedStyle(name="cd3", number_format="MMMM-YY")
# now that the style has been set up for Excel Writer we can  write the file.
Unit_5_Lighting_meter.to_excel(excel_writer=ewb, sheet_name="Unit_5_Light")
# after this we have to format the Unit_5_SmallPower_meter date column using the NamedStyles set up
ws = ewb.book["Unit_5_Light"]
for i in range(1, len(Unit_5_Lighting_meter) + 2):
    ws.cell(row=i, column=2).style = nsmmyy
    # ws.cell(row=i, column=3).style = nsmmmyy
    # ws.cell(row=i, column=2).style = nsbyy
ewb.close()

# Finally we need to sort out the Unit_5_SmallPower_meter column width value in the completed Excel sheet
worksheet = openpyxl.load_workbook(base_dest_path + "/" + "Unit_5_Lighting_meter.xlsx")
sheet = worksheet.active
sheet.column_dimensions['B'].width = 19
worksheet.save(base_dest_path + "/" + "Unit_5_Lighting_meter.xlsx")
worksheet.close()

# next meter

Admin_Build_Light_meter = pd.read_csv(base_path + "/" + file_name28, index_col=False, skiprows=[0])
Admin_Build_Light_meter.columns = column_names
Admin_Build_Light_meter['A'] = pd.to_datetime(Admin_Build_Light_meter['A'], unit='D', origin='1899-12-30')

Admin_Build_Light_meter.rename(columns={"A": "Admin_Light_Mtr"}, inplace=True)
Admin_Build_Light_meter['Total_Day'] = Admin_Build_Light_meter['24:00'] - Admin_Build_Light_meter['00:15']

# now let's write it as xlsx file format
# first we need to set up some format styles to be applied to the Excel sheet date column using the named_style module
ewb = pd.ExcelWriter(base_dest_path + "/" + "Admin_Building_Light_meter.xlsx", engine="openpyxl")
nsmmyy = NamedStyle(name="cd1", number_format="DD-MM-YY")
nsmmmyy = NamedStyle(name="cd2", number_format="MMM-YY")
nsbyy = NamedStyle(name="cd3", number_format="MMMM-YY")
# now that the style has been set up for Excel Writer we can  write the file.
Admin_Build_Light_meter.to_excel(excel_writer=ewb, sheet_name="Admin_Light_mtr")
# after this we have to format the Admin_Building_Light_meter date column using the NamedStyles set up
ws = ewb.book["Admin_Light_mtr"]
for i in range(1, len(Admin_Build_Light_meter) + 2):
    ws.cell(row=i, column=2).style = nsmmyy
    # ws.cell(row=i, column=3).style = nsmmmyy
    # ws.cell(row=i, column=2).style = nsbyy
ewb.close()

# Finally we need to sort out the Admin_Building_Light_meter column width value in the completed Excel sheet
worksheet = openpyxl.load_workbook(base_dest_path + "/" + "Admin_Building_Light_meter.xlsx")
sheet = worksheet.active
sheet.column_dimensions['B'].width = 19
worksheet.save(base_dest_path + "/" + "Admin_Building_Light_meter.xlsx")
worksheet.close()

# next meter

Admin_Build_GS_meter = pd.read_csv(base_path + "/" + file_name29, index_col=False, skiprows=[0])
Admin_Build_GS_meter.columns = column_names
Admin_Build_GS_meter['A'] = pd.to_datetime(Admin_Build_GS_meter['A'], unit='D', origin='1899-12-30')

Admin_Build_GS_meter.rename(columns={"A": "Unit_10_GS"}, inplace=True)
Admin_Build_GS_meter['Total_Day'] = Admin_Build_GS_meter['24:00'] - Admin_Build_GS_meter['00:15']

# now let's write it as xlsx file format
# first we need to set up some format styles to be applied to the Excel sheet date column using the named_style module
ewb = pd.ExcelWriter(base_dest_path + "/" + "Admin_Building_GS_meter.xlsx", engine="openpyxl")
nsmmyy = NamedStyle(name="cd1", number_format="DD-MM-YY")
nsmmmyy = NamedStyle(name="cd2", number_format="MMM-YY")
nsbyy = NamedStyle(name="cd3", number_format="MMMM-YY")
# now that the style has been set up for Excel Writer we can  write the file.
Admin_Build_GS_meter.to_excel(excel_writer=ewb, sheet_name="Admin_Gs_mtr")
# after this we have to format the Admin_Building_GS_meter date column using the NamedStyles set up
ws = ewb.book["Admin_Gs_mtr"]
for i in range(1, len(Admin_Build_GS_meter) + 2):
    ws.cell(row=i, column=2).style = nsmmyy
    # ws.cell(row=i, column=3).style = nsmmmyy
    # ws.cell(row=i, column=2).style = nsbyy
ewb.close()

# Finally we need to sort out the Admin_Building_GS_meter column width value in the completed Excel sheet
worksheet = openpyxl.load_workbook(base_dest_path + "/" + "Admin_Building_GS_meter.xlsx")
sheet = worksheet.active
sheet.column_dimensions['B'].width = 19
worksheet.save(base_dest_path + "/" + "Admin_Building_GS_meter.xlsx")
worksheet.close()

# next meter

Admin_Build_SmallPwr_meter = pd.read_csv(base_path + "/" + file_name30, index_col=False, skiprows=[0])
Admin_Build_SmallPwr_meter.columns = column_names
Admin_Build_SmallPwr_meter['A'] = pd.to_datetime(Admin_Build_SmallPwr_meter['A'], unit='D', origin='1899-12-30')

Admin_Build_SmallPwr_meter.rename(columns={"A": "Admin_Build_SmlPwr"}, inplace=True)
Admin_Build_SmallPwr_meter['Total_Day'] = Admin_Build_SmallPwr_meter['24:00'] - Admin_Build_SmallPwr_meter['00:15']

# now let's write it as xlsx file format
# first we need to set up some format styles to be applied to the Excel sheet date column using the named_style module
ewb = pd.ExcelWriter(base_dest_path + "/" + "Admin_Building_SmallPower_meter.xlsx", engine="openpyxl")
nsmmyy = NamedStyle(name="cd1", number_format="DD-MM-YY")
nsmmmyy = NamedStyle(name="cd2", number_format="MMM-YY")
nsbyy = NamedStyle(name="cd3", number_format="MMMM-YY")
# now that the style has been set up for Excel Writer we can  write the file.
Admin_Build_SmallPwr_meter.to_excel(excel_writer=ewb, sheet_name="Admin_SmlPwr_mtr")
# after this we have to format the Admin_Building_SmlPwr_meter date column using the NamedStyles set up
ws = ewb.book["Admin_SmlPwr_mtr"]
for i in range(1, len(Admin_Build_SmallPwr_meter) + 2):
    ws.cell(row=i, column=2).style = nsmmyy
    # ws.cell(row=i, column=3).style = nsmmmyy
    # ws.cell(row=i, column=2).style = nsbyy
ewb.close()

# Finally we need to sort out the Admin_Building_SmlPwr_meter column width value in the completed Excel sheet
worksheet = openpyxl.load_workbook(base_dest_path + "/" + "Admin_Building_SmallPower_meter.xlsx")
sheet = worksheet.active
sheet.column_dimensions['B'].width = 19
worksheet.save(base_dest_path + "/" + "Admin_Building_SmallPower_meter.xlsx")
worksheet.close()

# next meter

Admin_Building_Water_meter = pd.read_csv(base_path + "/" + file_name31, index_col=False, skiprows=[0])
Admin_Building_Water_meter.columns = column_names
Admin_Building_Water_meter['A'] = pd.to_datetime(Admin_Building_Water_meter['A'], unit='D', origin='1899-12-30')

Admin_Building_Water_meter.rename(columns={"A": "Admin_Build_Wtr_Mtr"}, inplace=True)
Admin_Building_Water_meter['Total_Day'] = Admin_Building_Water_meter['24:00'] - Admin_Building_Water_meter['00:15']

# now let's write it as xlsx file format
# first we need to set up some format styles to be applied to the Excel sheet date column using the named_style module
ewb = pd.ExcelWriter(base_dest_path + "/" + "Admin_Building_Water_meter.xlsx", engine="openpyxl")
nsmmyy = NamedStyle(name="cd1", number_format="DD-MM-YY")
nsmmmyy = NamedStyle(name="cd2", number_format="MMM-YY")
nsbyy = NamedStyle(name="cd3", number_format="MMMM-YY")
# now that the style has been set up for Excel Writer we can  write the file.
Admin_Building_Water_meter.to_excel(excel_writer=ewb, sheet_name="Admin_Water_mtr")
# after this we have to format the Admin_Building_water_meter date column using the NamedStyles set up
ws = ewb.book["Admin_Water_mtr"]
for i in range(1, len(Admin_Building_Water_meter) + 2):
    ws.cell(row=i, column=2).style = nsmmyy
    # ws.cell(row=i, column=3).style = nsmmmyy
    # ws.cell(row=i, column=2).style = nsbyy
ewb.close()

# Finally we need to sort out the Admin_Building_water_meter column width value in the completed Excel sheet
worksheet = openpyxl.load_workbook(base_dest_path + "/" + "Admin_Building_Water_meter.xlsx")
sheet = worksheet.active
sheet.column_dimensions['B'].width = 19
worksheet.save(base_dest_path + "/" + "Admin_Building_Water_meter.xlsx")
worksheet.close()

# next meter

Admin_Building_Gas_meter = pd.read_csv(base_path + "/" + file_name32, index_col=False, skiprows=[0])
Admin_Building_Gas_meter.columns = column_names
Admin_Building_Gas_meter['A'] = pd.to_datetime(Admin_Building_Gas_meter['A'], unit='D', origin='1899-12-30')

Admin_Building_Gas_meter.rename(columns={"A": "Admin_Build_Wtr_Mtr"}, inplace=True)
Admin_Building_Gas_meter['Total_Day'] = Admin_Building_Gas_meter['24:00'] - Admin_Building_Gas_meter['00:15']

# now let's write it as xlsx file format
# first we need to set up some format styles to be applied to the Excel sheet date column using the named_style module
ewb = pd.ExcelWriter(base_dest_path + "/" + "Admin_Building_Gas_meter.xlsx", engine="openpyxl")
nsmmyy = NamedStyle(name="cd1", number_format="DD-MM-YY")
nsmmmyy = NamedStyle(name="cd2", number_format="MMM-YY")
nsbyy = NamedStyle(name="cd3", number_format="MMMM-YY")
# now that the style has been set up for Excel Writer we can  write the file.
Admin_Building_Gas_meter.to_excel(excel_writer=ewb, sheet_name="Admin_Gas_mtr")
# after this we have to format the Admin_Building_Gas_meter date column using the NamedStyles set up
ws = ewb.book["Admin_Gas_mtr"]
for i in range(1, len(Admin_Building_Gas_meter) + 2):
    ws.cell(row=i, column=2).style = nsmmyy
    # ws.cell(row=i, column=3).style = nsmmmyy
    # ws.cell(row=i, column=2).style = nsbyy
ewb.close()

# Finally we need to sort out the Admin_Building_Gas_meter column width value in the completed Excel sheet
worksheet = openpyxl.load_workbook(base_dest_path + "/" + "Admin_Building_Gas_meter.xlsx")
sheet = worksheet.active
sheet.column_dimensions['B'].width = 19
worksheet.save(base_dest_path + "/" + "Admin_Building_Gas_meter.xlsx")
worksheet.close()

# next meter

EdRec_Lighting_meter = pd.read_csv(base_path + "/" + file_name33, index_col=False, skiprows=[0])
EdRec_Lighting_meter.columns = column_names
EdRec_Lighting_meter['A'] = pd.to_datetime(EdRec_Lighting_meter['A'], unit='D', origin='1899-12-30')

EdRec_Lighting_meter.rename(columns={"A": "Edrec_Light_Mtr"}, inplace=True)
EdRec_Lighting_meter['Total_Day'] = EdRec_Lighting_meter['24:00'] - EdRec_Lighting_meter['00:15']

# now let's write it as xlsx file format
# first we need to set up some format styles to be applied to the Excel sheet date column using the named_style module
ewb = pd.ExcelWriter(base_dest_path + "/" + "EdRec_Lighting_meter.xlsx", engine="openpyxl")
nsmmyy = NamedStyle(name="cd1", number_format="DD-MM-YY")
nsmmmyy = NamedStyle(name="cd2", number_format="MMM-YY")
nsbyy = NamedStyle(name="cd3", number_format="MMMM-YY")
# now that the style has been set up for Excel Writer we can  write the file.
EdRec_Lighting_meter.to_excel(excel_writer=ewb, sheet_name="EdRec_Light_meter")
# after this we have to format the EdRec_Lighting_meter date column using the NamedStyles set up
ws = ewb.book["EdRec_Light_meter"]
for i in range(1, len(EdRec_Lighting_meter) + 2):
    ws.cell(row=i, column=2).style = nsmmyy
    # ws.cell(row=i, column=3).style = nsmmmyy
    # ws.cell(row=i, column=2).style = nsbyy
ewb.close()

# Finally we need to sort out the EdRec_Lighting_meter column width value in the completed Excel sheet
worksheet = openpyxl.load_workbook(base_dest_path + "/" + "EdRec_Lighting_meter.xlsx")
sheet = worksheet.active
sheet.column_dimensions['B'].width = 19
worksheet.save(base_dest_path + "/" + "EdRec_Lighting_meter.xlsx")
worksheet.close()

# next meter

EdRec_GS_meter = pd.read_csv(base_path + "/" + file_name34, index_col=False, skiprows=[0])
EdRec_GS_meter.columns = column_names
EdRec_GS_meter['A'] = pd.to_datetime(EdRec_GS_meter['A'], unit='D', origin='1899-12-30')

EdRec_GS_meter.rename(columns={"A": "Edrec_GS_Mtr"}, inplace=True)
EdRec_GS_meter['Total_Day'] = EdRec_GS_meter['24:00'] - EdRec_GS_meter['00:15']

# now let's write it as xlsx file format
# first we need to set up some format styles to be applied to the Excel sheet date column using the named_style module
ewb = pd.ExcelWriter(base_dest_path + "/" + "EdRec_GS_meter.xlsx", engine="openpyxl")
nsmmyy = NamedStyle(name="cd1", number_format="DD-MM-YY")
nsmmmyy = NamedStyle(name="cd2", number_format="MMM-YY")
nsbyy = NamedStyle(name="cd3", number_format="MMMM-YY")
# now that the style has been set up for Excel Writer we can  write the file.
EdRec_GS_meter.to_excel(excel_writer=ewb, sheet_name="EdRec_Gs_meter")
# after this we have to format the EdRec_Gs_meter date column using the NamedStyles set up
ws = ewb.book["EdRec_Gs_meter"]
for i in range(1, len(EdRec_GS_meter) + 2):
    ws.cell(row=i, column=2).style = nsmmyy
    # ws.cell(row=i, column=3).style = nsmmmyy
    # ws.cell(row=i, column=2).style = nsbyy
ewb.close()

# Finally we need to sort out the EdRec_Gs_meter column width value in the completed Excel sheet
worksheet = openpyxl.load_workbook(base_dest_path + "/" + "EdRec_GS_meter.xlsx")
sheet = worksheet.active
sheet.column_dimensions['B'].width = 19
worksheet.save(base_dest_path + "/" + "EdRec_GS_meter.xlsx")
worksheet.close()

# next meter

EdRec_SmallPower_meter = pd.read_csv(base_path + "/" + file_name35, index_col=False, skiprows=[0])
EdRec_SmallPower_meter.columns = column_names
EdRec_SmallPower_meter['A'] = pd.to_datetime(EdRec_SmallPower_meter['A'], unit='D', origin='1899-12-30')

EdRec_SmallPower_meter.rename(columns={"A": "Edrec_SmlPwr_Mtr"}, inplace=True)
EdRec_SmallPower_meter['Total_Day'] = EdRec_SmallPower_meter['24:00'] - EdRec_SmallPower_meter['00:15']

# now let's write it as xlsx file format
# first we need to set up some format styles to be applied to the Excel sheet date column using the named_style module
ewb = pd.ExcelWriter(base_dest_path + "/" + "EdRec_SmallPower_meter.xlsx", engine="openpyxl")
nsmmyy = NamedStyle(name="cd1", number_format="DD-MM-YY")
nsmmmyy = NamedStyle(name="cd2", number_format="MMM-YY")
nsbyy = NamedStyle(name="cd3", number_format="MMMM-YY")
# now that the style has been set up for Excel Writer we can  write the file.
EdRec_SmallPower_meter.to_excel(excel_writer=ewb, sheet_name="EdRec_SmlPwr_meter")
# after this we have to format the EdRec_SmallPower_meter date column using the NamedStyles set up
ws = ewb.book["EdRec_SmlPwr_meter"]
for i in range(1, len(EdRec_SmallPower_meter) + 2):
    ws.cell(row=i, column=2).style = nsmmyy
    # ws.cell(row=i, column=3).style = nsmmmyy
    # ws.cell(row=i, column=2).style = nsbyy
ewb.close()

# Finally we need to sort out the EdRec_SmallPower_meter column width value in the completed Excel sheet
worksheet = openpyxl.load_workbook(base_dest_path + "/" + "EdRec_SmallPower_meter.xlsx")
sheet = worksheet.active
sheet.column_dimensions['B'].width = 19
worksheet.save(base_dest_path + "/" + "EdRec_SmallPower_meter.xlsx")
worksheet.close()

# next meter

EdRec_Water_meter = pd.read_csv(base_path + "/" + file_name36, index_col=False, skiprows=[0])
EdRec_Water_meter.columns = column_names
EdRec_Water_meter['A'] = pd.to_datetime(EdRec_Water_meter['A'], unit='D', origin='1899-12-30')

EdRec_Water_meter.rename(columns={"A": "Edrec_Wtr_Mtr"}, inplace=True)
EdRec_Water_meter['Total_Day'] = EdRec_Water_meter['24:00'] - EdRec_Water_meter['00:15']

# now let's write it as xlsx file format
# first we need to set up some format styles to be applied to the Excel sheet date column using the named_style module
ewb = pd.ExcelWriter(base_dest_path + "/" + "EdRec_Water_meter.xlsx", engine="openpyxl")
nsmmyy = NamedStyle(name="cd1", number_format="DD-MM-YY")
nsmmmyy = NamedStyle(name="cd2", number_format="MMM-YY")
nsbyy = NamedStyle(name="cd3", number_format="MMMM-YY")
# now that the style has been set up for Excel Writer we can  write the file.
EdRec_Water_meter.to_excel(excel_writer=ewb, sheet_name="EdRec_Water_meter")
# after this we have to format the EdRec_Water_meter date column using the NamedStyles set up
ws = ewb.book["EdRec_Water_meter"]
for i in range(1, len(EdRec_Water_meter) + 2):
    ws.cell(row=i, column=2).style = nsmmyy
    # ws.cell(row=i, column=3).style = nsmmmyy
    # ws.cell(row=i, column=2).style = nsbyy
ewb.close()

# Finally we need to sort out the EdRec_Water_meter column width value in the completed Excel sheet
worksheet = openpyxl.load_workbook(base_dest_path + "/" + "EdRec_Water_meter.xlsx")
sheet = worksheet.active
sheet.column_dimensions['B'].width = 19
worksheet.save(base_dest_path + "/" + "EdRec_Water_meter.xlsx")
worksheet.close()

# next meter

Unit_4_Lighting_meter = pd.read_csv(base_path + "/" + file_name37, index_col=False, skiprows=[0])
Unit_4_Lighting_meter.columns = column_names
Unit_4_Lighting_meter['A'] = pd.to_datetime(Unit_4_Lighting_meter['A'], unit='D', origin='1899-12-30')

Unit_4_Lighting_meter.rename(columns={"A": "Unit_4_Light"}, inplace=True)
Unit_4_Lighting_meter['Total_Day'] = Unit_4_Lighting_meter['24:00'] - Unit_4_Lighting_meter['00:15']

# now let's write it as xlsx file format
# first we need to set up some format styles to be applied to the Excel sheet date column using the named_style module
ewb = pd.ExcelWriter(base_dest_path + "/" + "Unit_4_Lighting_meter.xlsx", engine="openpyxl")
nsmmyy = NamedStyle(name="cd1", number_format="DD-MM-YY")
nsmmmyy = NamedStyle(name="cd2", number_format="MMM-YY")
nsbyy = NamedStyle(name="cd3", number_format="MMMM-YY")
# now that the style has been set up for Excel Writer we can  write the file.
Unit_4_Lighting_meter.to_excel(excel_writer=ewb, sheet_name="Unit_4_Light_mtr")
# after this we have to format the Unit_4_Lighting_meter date column using the NamedStyles set up
ws = ewb.book["Unit_4_Light_mtr"]
for i in range(1, len(Unit_4_Lighting_meter) + 2):
    ws.cell(row=i, column=2).style = nsmmyy
    # ws.cell(row=i, column=3).style = nsmmmyy
    # ws.cell(row=i, column=2).style = nsbyy
ewb.close()

# Finally we need to sort out the Unit_4_Lighting_meter column width value in the completed Excel sheet
worksheet = openpyxl.load_workbook(base_dest_path + "/" + "Unit_4_Lighting_meter.xlsx")
sheet = worksheet.active
sheet.column_dimensions['B'].width = 19
worksheet.save(base_dest_path + "/" + "Unit_4_Lighting_meter.xlsx")
worksheet.close()

# next meter

Unit_4_GS_meter = pd.read_csv(base_path + "/" + file_name38, index_col=False, skiprows=[0])
Unit_4_GS_meter.columns = column_names
Unit_4_GS_meter['A'] = pd.to_datetime(Unit_4_GS_meter['A'], unit='D', origin='1899-12-30')

Unit_4_GS_meter.rename(columns={"A": "Unit_4_GS"}, inplace=True)
Unit_4_GS_meter['Total_Day'] = Unit_4_GS_meter['24:00'] - Unit_4_GS_meter['00:15']

# now let's write it as xlsx file format
# first we need to set up some format styles to be applied to the Excel sheet date column using the named_style module
ewb = pd.ExcelWriter(base_dest_path + "/" + "Unit_4_GS_meter.xlsx", engine="openpyxl")
nsmmyy = NamedStyle(name="cd1", number_format="DD-MM-YY")
nsmmmyy = NamedStyle(name="cd2", number_format="MMM-YY")
nsbyy = NamedStyle(name="cd3", number_format="MMMM-YY")
# now that the style has been set up for Excel Writer we can  write the file.
Unit_4_GS_meter.to_excel(excel_writer=ewb, sheet_name="Unit_4_GS_mtr")
# after this we have to format theUnit_4_GS_meter date column using the NamedStyles set up
ws = ewb.book["Unit_4_GS_mtr"]
for i in range(1, len(Unit_4_GS_meter) + 2):
    ws.cell(row=i, column=2).style = nsmmyy
    # ws.cell(row=i, column=3).style = nsmmmyy
    # ws.cell(row=i, column=2).style = nsbyy
ewb.close()

# Finally we need to sort out the Unit_4_GS_meter column width value in the completed Excel sheet
worksheet = openpyxl.load_workbook(base_dest_path + "/" + "Unit_4_GS_meter.xlsx")
sheet = worksheet.active
sheet.column_dimensions['B'].width = 19
worksheet.save(base_dest_path + "/" + "Unit_4_GS_meter.xlsx")
worksheet.close()

# next meter

Unit_4_Water_meter = pd.read_csv(base_path + "/" + file_name39, index_col=False, skiprows=[0])
Unit_4_Water_meter.columns = column_names
Unit_4_Water_meter['A'] = pd.to_datetime(Unit_4_Water_meter['A'], unit='D', origin='1899-12-30')

Unit_4_Water_meter.rename(columns={"A": "Unit_4_Wtr"}, inplace=True)
Unit_4_Water_meter['Total_Day'] = Unit_4_Water_meter['24:00'] - Unit_4_Water_meter['00:15']

# now let's write it as xlsx file format
# first we need to set up some format styles to be applied to the Excel sheet date column using the named_style module
ewb = pd.ExcelWriter(base_dest_path + "/" + "Unit_4_Water_meter.xlsx", engine="openpyxl")
nsmmyy = NamedStyle(name="cd1", number_format="DD-MM-YY")
nsmmmyy = NamedStyle(name="cd2", number_format="MMM-YY")
nsbyy = NamedStyle(name="cd3", number_format="MMMM-YY")
# now that the style has been set up for Excel Writer we can  write the file.
Unit_4_Water_meter.to_excel(excel_writer=ewb, sheet_name="Unit_4_Water_mtr")
# after this we have to format the Unit_4_Water_meter date column using the NamedStyles set up
ws = ewb.book["Unit_4_Water_mtr"]
for i in range(1, len(Unit_4_Water_meter) + 2):
    ws.cell(row=i, column=2).style = nsmmyy
    # ws.cell(row=i, column=3).style = nsmmmyy
    # ws.cell(row=i, column=2).style = nsbyy
ewb.close()

# Finally we need to sort out the Unit_4_Water_meter column width value in the completed Excel sheet
worksheet = openpyxl.load_workbook(base_dest_path + "/" + "Unit_4_Water_meter.xlsx")
sheet = worksheet.active
sheet.column_dimensions['B'].width = 19
worksheet.save(base_dest_path + "/" + "Unit_4_Water_meter.xlsx")
worksheet.close()

# next meter

Unit_4_SmallPower_meter = pd.read_csv(base_path + "/" + file_name40, index_col=False, skiprows=[0])
Unit_4_SmallPower_meter.columns = column_names
Unit_4_SmallPower_meter['A'] = pd.to_datetime(Unit_4_SmallPower_meter['A'], unit='D', origin='1899-12-30')

Unit_4_SmallPower_meter.rename(columns={"A": "Unit_4_SmlPwr"}, inplace=True)
Unit_4_SmallPower_meter['Total_Day'] = Unit_4_SmallPower_meter['24:00'] - Unit_4_SmallPower_meter['00:15']

# now let's write it as xlsx file format
# first we need to set up some format styles to be applied to the Excel sheet date column using the named_style module
ewb = pd.ExcelWriter(base_dest_path + "/" + "Unit_4_SmallPower_meter.xlsx", engine="openpyxl")
nsmmyy = NamedStyle(name="cd1", number_format="DD-MM-YY")
nsmmmyy = NamedStyle(name="cd2", number_format="MMM-YY")
nsbyy = NamedStyle(name="cd3", number_format="MMMM-YY")
# now that the style has been set up for Excel Writer we can  write the file.
Unit_4_SmallPower_meter.to_excel(excel_writer=ewb, sheet_name="Unit_4_SmlPwr_mtr")
# after this we have to format the Unit_4_SmallPower_meter date column using the NamedStyles set up
ws = ewb.book["Unit_4_SmlPwr_mtr"]
for i in range(1, len(Unit_4_SmallPower_meter) + 2):
    ws.cell(row=i, column=2).style = nsmmyy
    # ws.cell(row=i, column=3).style = nsmmmyy
    # ws.cell(row=i, column=2).style = nsbyy
ewb.close()

# Finally we need to sort out the Unit_4_SmallPower_meter column width value in the completed Excel sheet
worksheet = openpyxl.load_workbook(base_dest_path + "/" + "Unit_4_SmallPower_meter.xlsx")
sheet = worksheet.active
sheet.column_dimensions['B'].width = 19
worksheet.save(base_dest_path + "/" + "Unit_4_SmallPower_meter.xlsx")
worksheet.close()

# next meter

Work_Shop_Lighting_meter = pd.read_csv(base_path + "/" + file_name41, index_col=False, skiprows=[0])
Work_Shop_Lighting_meter.columns = column_names
Work_Shop_Lighting_meter['A'] = pd.to_datetime(Work_Shop_Lighting_meter['A'], unit='D', origin='1899-12-30')

Work_Shop_Lighting_meter.rename(columns={"A": "Wrkshp_Light_Mtr"}, inplace=True)
Work_Shop_Lighting_meter['Total_Day'] = Work_Shop_Lighting_meter['24:00'] - Work_Shop_Lighting_meter['00:15']

# now let's write it as xlsx file format
# first we need to set up some format styles to be applied to the Excel sheet date column using the named_style module
ewb = pd.ExcelWriter(base_dest_path + "/" + "Work_Shop_Lighting_meter.xlsx", engine="openpyxl")
nsmmyy = NamedStyle(name="cd1", number_format="DD-MM-YY")
nsmmmyy = NamedStyle(name="cd2", number_format="MMM-YY")
nsbyy = NamedStyle(name="cd3", number_format="MMMM-YY")
# now that the style has been set up for Excel Writer we can  write the file.
Work_Shop_Lighting_meter.to_excel(excel_writer=ewb, sheet_name="Work_Shop_Light_mtr")
# after this we have to format the Work_Shop_Lighting_meter date column using the NamedStyles set up
ws = ewb.book["Work_Shop_Light_mtr"]
for i in range(1, len(Work_Shop_Lighting_meter) + 2):
    ws.cell(row=i, column=2).style = nsmmyy
    # ws.cell(row=i, column=3).style = nsmmmyy
    # ws.cell(row=i, column=2).style = nsbyy
ewb.close()

# Finally we need to sort out the Work_Shop_Lighting_meter column width value in the completed Excel sheet
worksheet = openpyxl.load_workbook(base_dest_path + "/" + "Work_Shop_Lighting_meter.xlsx")
sheet = worksheet.active
sheet.column_dimensions['B'].width = 19
worksheet.save(base_dest_path + "/" + "Work_Shop_Lighting_meter.xlsx")
worksheet.close()

# next meter

Work_Shop_GS_meter = pd.read_csv(base_path + "/" + file_name42, index_col=False, skiprows=[0])
Work_Shop_GS_meter.columns = column_names
Work_Shop_GS_meter['A'] = pd.to_datetime(Work_Shop_GS_meter['A'], unit='D', origin='1899-12-30')

Work_Shop_GS_meter.rename(columns={"A": "Wrkshp_GS_Mtr"}, inplace=True)
Work_Shop_GS_meter['Total_Day'] = Work_Shop_GS_meter['24:00'] - Work_Shop_GS_meter['00:15']

# now let's write it as xlsx file format
# first we need to set up some format styles to be applied to the Excel sheet date column using the named_style module
ewb = pd.ExcelWriter(base_dest_path + "/" + "Work_Shop_GS_meter.xlsx", engine="openpyxl")
nsmmyy = NamedStyle(name="cd1", number_format="DD-MM-YY")
nsmmmyy = NamedStyle(name="cd2", number_format="MMM-YY")
nsbyy = NamedStyle(name="cd3", number_format="MMMM-YY")
# now that the style has been set up for Excel Writer we can  write the file.
Work_Shop_GS_meter.to_excel(excel_writer=ewb, sheet_name="Work_Shop_GS_mtr")
# after this we have to format the Work_Shop_Lighting_meter date column using the NamedStyles set up
ws = ewb.book["Work_Shop_GS_mtr"]
for i in range(1, len(Work_Shop_GS_meter) + 2):
    ws.cell(row=i, column=2).style = nsmmyy
    # ws.cell(row=i, column=3).style = nsmmmyy
    # ws.cell(row=i, column=2).style = nsbyy
ewb.close()

# Finally we need to sort out the Work_Shop_GS_meter column width value in the completed Excel sheet
worksheet = openpyxl.load_workbook(base_dest_path + "/" + "Work_Shop_GS_meter.xlsx")
sheet = worksheet.active
sheet.column_dimensions['B'].width = 19
worksheet.save(base_dest_path + "/" + "Work_Shop_GS_meter.xlsx")
worksheet.close()

# next meter

Eng_Cntr_Ext_Svc_meter = pd.read_csv(base_path + "/" + file_name43, index_col=False, skiprows=[0])
Eng_Cntr_Ext_Svc_meter.columns = column_names
Eng_Cntr_Ext_Svc_meter['A'] = pd.to_datetime(Eng_Cntr_Ext_Svc_meter['A'], unit='D', origin='1899-12-30')

Eng_Cntr_Ext_Svc_meter.rename(columns={"A": "Eng_Cntr_ExtSvc"}, inplace=True)
Eng_Cntr_Ext_Svc_meter['Total_Day'] = Eng_Cntr_Ext_Svc_meter['24:00'] - Eng_Cntr_Ext_Svc_meter['00:15']

# now let's write it as xlsx file format
# first we need to set up some format styles to be applied to the Excel sheet date column using the named_style module
ewb = pd.ExcelWriter(base_dest_path + "/" + "Eng_Cntr_Ext_Svc_meter.xlsx", engine="openpyxl")
nsmmyy = NamedStyle(name="cd1", number_format="DD-MM-YY")
nsmmmyy = NamedStyle(name="cd2", number_format="MMM-YY")
nsbyy = NamedStyle(name="cd3", number_format="MMMM-YY")
# now that the style has been set up for Excel Writer we can  write the file.
Eng_Cntr_Ext_Svc_meter.to_excel(excel_writer=ewb, sheet_name="Eng_Cntr_Ext_Svc_mtr")
# after this we have to format the Eng_Cntr_Ext_Svc_mtr date column using the NamedStyles set up
ws = ewb.book["Eng_Cntr_Ext_Svc_mtr"]
for i in range(1, len(Eng_Cntr_Ext_Svc_meter) + 2):
    ws.cell(row=i, column=2).style = nsmmyy
    # ws.cell(row=i, column=3).style = nsmmmyy
    # ws.cell(row=i, column=2).style = nsbyy
ewb.close()

# Finally we need to sort out the Eng_Cntr_Ext_Svc_meter column width value in the completed Excel sheet
worksheet = openpyxl.load_workbook(base_dest_path + "/" + "Eng_Cntr_Ext_Svc_meter.xlsx")
sheet = worksheet.active
sheet.column_dimensions['B'].width = 19
worksheet.save(base_dest_path + "/" + "Eng_Cntr_Ext_Svc_meter.xlsx")
worksheet.close()

# next meter

Eng_Cntr_Comms_Rm_meter = pd.read_csv(base_path + "/" + file_name44, index_col=False, skiprows=[0])
Eng_Cntr_Comms_Rm_meter.columns = column_names
Eng_Cntr_Comms_Rm_meter['A'] = pd.to_datetime(Eng_Cntr_Comms_Rm_meter['A'], unit='D', origin='1899-12-30')

Eng_Cntr_Comms_Rm_meter.rename(columns={"A": "Eng_Cntr_Coms-Rm"}, inplace=True)
Eng_Cntr_Comms_Rm_meter['Total_Day'] = Eng_Cntr_Comms_Rm_meter['24:00'] - Eng_Cntr_Comms_Rm_meter['00:15']

# now let's write it as xlsx file format
# first we need to set up some format styles to be applied to the Excel sheet date column using the named_style module
ewb = pd.ExcelWriter(base_dest_path + "/" + "Eng_Cntr_Comms_Rm_meter.xlsx", engine="openpyxl")
nsmmyy = NamedStyle(name="cd1", number_format="DD-MM-YY")
nsmmmyy = NamedStyle(name="cd2", number_format="MMM-YY")
nsbyy = NamedStyle(name="cd3", number_format="MMMM-YY")
# now that the style has been set up for Excel Writer we can  write the file.
Eng_Cntr_Comms_Rm_meter.to_excel(excel_writer=ewb, sheet_name="Eng_Cntr_Comms_Rm_mtr")
# after this we have to format the Eng_Cntr_Comms_Rm_mtr date column using the NamedStyles set up
ws = ewb.book["Eng_Cntr_Comms_Rm_mtr"]
for i in range(1, len(Eng_Cntr_Comms_Rm_meter) + 2):
    ws.cell(row=i, column=2).style = nsmmyy
    # ws.cell(row=i, column=3).style = nsmmmyy
    # ws.cell(row=i, column=2).style = nsbyy
ewb.close()

# Finally we need to sort out the Work_Shop_Lighting_meter column width value in the completed Excel sheet
worksheet = openpyxl.load_workbook(base_dest_path + "/" + "Eng_Cntr_Comms_Rm_meter.xlsx")
sheet = worksheet.active
sheet.column_dimensions['B'].width = 19
worksheet.save(base_dest_path + "/" + "Eng_Cntr_Comms_Rm_meter.xlsx")
worksheet.close()

# next meter

Eng_Cntr_Admin_Bld_meter = pd.read_csv(base_path + "/" + file_name45, index_col=False, skiprows=[0])
Eng_Cntr_Admin_Bld_meter.columns = column_names
Eng_Cntr_Admin_Bld_meter['A'] = pd.to_datetime(Eng_Cntr_Admin_Bld_meter['A'], unit='D', origin='1899-12-30')

Eng_Cntr_Admin_Bld_meter.rename(columns={"A": "Eng_Cntr_Admin_Bld"}, inplace=True)
Eng_Cntr_Admin_Bld_meter['Total_Day'] = Eng_Cntr_Admin_Bld_meter['24:00'] - Eng_Cntr_Admin_Bld_meter['00:15']

# now let's write it as xlsx file format
# first we need to set up some format styles to be applied to the Excel sheet date column using the named_style module
ewb = pd.ExcelWriter(base_dest_path + "/" + "Eng_Cntr_Admin_Bld_meter.xlsx", engine="openpyxl")
nsmmyy = NamedStyle(name="cd1", number_format="DD-MM-YY")
nsmmmyy = NamedStyle(name="cd2", number_format="MMM-YY")
nsbyy = NamedStyle(name="cd3", number_format="MMMM-YY")
# now that the style has been set up for Excel Writer we can  write the file.
Eng_Cntr_Admin_Bld_meter.to_excel(excel_writer=ewb, sheet_name="Eng_Cntr_Admin_Bld_mtr")
# after this we have to format the Eng_Cntr_Admin_Bld_meter date column using the NamedStyles set up
ws = ewb.book["Eng_Cntr_Admin_Bld_mtr"]
for i in range(1, len(Eng_Cntr_Admin_Bld_meter) + 2):
    ws.cell(row=i, column=2).style = nsmmyy
    # ws.cell(row=i, column=3).style = nsmmmyy
    # ws.cell(row=i, column=2).style = nsbyy
ewb.close()

# Finally we need to sort out the Eng_Cntr_Admin_Bld_meter column width value in the completed Excel sheet
worksheet = openpyxl.load_workbook(base_dest_path + "/" + "Eng_Cntr_Admin_Bld_meter.xlsx")
sheet = worksheet.active
sheet.column_dimensions['B'].width = 19
worksheet.save(base_dest_path + "/" + "Eng_Cntr_Admin_Bld_meter.xlsx")
worksheet.close()

# next meter

Eng_Cntr_Sec_Suite_meter = pd.read_csv(base_path + "/" + file_name46, index_col=False, skiprows=[0])
Eng_Cntr_Sec_Suite_meter.columns = column_names
Eng_Cntr_Sec_Suite_meter['A'] = pd.to_datetime(Eng_Cntr_Sec_Suite_meter['A'], unit='D', origin='1899-12-30')

Eng_Cntr_Sec_Suite_meter.rename(columns={"A": "Eng_Cntr_Sec_Suite"}, inplace=True)
Eng_Cntr_Sec_Suite_meter['Total_Day'] = Eng_Cntr_Sec_Suite_meter['24:00'] - Eng_Cntr_Sec_Suite_meter['00:15']

# now let's write it as xlsx file format
# first we need to set up some format styles to be applied to the Excel sheet date column using the named_style module
ewb = pd.ExcelWriter(base_dest_path + "/" + "Eng_Cntr_Sec_Suite_meter.xlsx", engine="openpyxl")
nsmmyy = NamedStyle(name="cd1", number_format="DD-MM-YY")
nsmmmyy = NamedStyle(name="cd2", number_format="MMM-YY")
nsbyy = NamedStyle(name="cd3", number_format="MMMM-YY")
# now that the style has been set up for Excel Writer we can  write the file.
Eng_Cntr_Sec_Suite_meter.to_excel(excel_writer=ewb, sheet_name="Eng_Cntr_Sec_Suite_mtr")
# after this we have to format the Eng_Cntr_Sec_Suite_meter date column using the NamedStyles set up
ws = ewb.book["Eng_Cntr_Sec_Suite_mtr"]
for i in range(1, len(Eng_Cntr_Sec_Suite_meter) + 2):
    ws.cell(row=i, column=2).style = nsmmyy
    # ws.cell(row=i, column=3).style = nsmmmyy
    # ws.cell(row=i, column=2).style = nsbyy
ewb.close()

# Finally we need to sort out the Eng_Cntr_Sec_Suite_meter column width value in the completed Excel sheet
worksheet = openpyxl.load_workbook(base_dest_path + "/" + "Eng_Cntr_Sec_Suite_meter.xlsx")
sheet = worksheet.active
sheet.column_dimensions['B'].width = 19
worksheet.save(base_dest_path + "/" + "Eng_Cntr_Sec_Suite_meter.xlsx")
worksheet.close()

# next meter

Eng_Cntr_Maint_Bld_meter = pd.read_csv(base_path + "/" + file_name47, index_col=False, skiprows=[0])
Eng_Cntr_Maint_Bld_meter.columns = column_names
Eng_Cntr_Maint_Bld_meter['A'] = pd.to_datetime(Eng_Cntr_Maint_Bld_meter['A'], unit='D', origin='1899-12-30')

Eng_Cntr_Maint_Bld_meter.rename(columns={"A": "Eng_Cntr_Maint_Bld"}, inplace=True)
Eng_Cntr_Maint_Bld_meter['Total_Day'] = Eng_Cntr_Maint_Bld_meter['24:00'] - Eng_Cntr_Maint_Bld_meter['00:15']

# now let's write it as xlsx file format
# first we need to set up some format styles to be applied to the Excel sheet date column using the named_style module
ewb = pd.ExcelWriter(base_dest_path + "/" + "Eng_Cntr_Maint_Bld_meter.xlsx", engine="openpyxl")
nsmmyy = NamedStyle(name="cd1", number_format="DD-MM-YY")
nsmmmyy = NamedStyle(name="cd2", number_format="MMM-YY")
nsbyy = NamedStyle(name="cd3", number_format="MMMM-YY")
# now that the style has been set up for Excel Writer we can  write the file.
Eng_Cntr_Maint_Bld_meter.to_excel(excel_writer=ewb, sheet_name="Eng_Cntr_Maint_Bld_mtr")
# after this we have to format the Eng_Cntr_Maint_Bld_meter date column using the NamedStyles set up
ws = ewb.book["Eng_Cntr_Maint_Bld_mtr"]
for i in range(1, len(Eng_Cntr_Maint_Bld_meter) + 2):
    ws.cell(row=i, column=2).style = nsmmyy
    # ws.cell(row=i, column=3).style = nsmmmyy
    # ws.cell(row=i, column=2).style = nsbyy
ewb.close()

# Finally we need to sort out the Eng_Cntr_Maint_Bld_meter column width value in the completed Excel sheet
worksheet = openpyxl.load_workbook(base_dest_path + "/" + "Eng_Cntr_Maint_Bld_meter.xlsx")
sheet = worksheet.active
sheet.column_dimensions['B'].width = 19
worksheet.save(base_dest_path + "/" + "Eng_Cntr_Maint_Bld_meter.xlsx")
worksheet.close()

# next meter

Eng_Cntr_Educ_Blk_meter = pd.read_csv(base_path + "/" + file_name48, index_col=False, skiprows=[0])
Eng_Cntr_Educ_Blk_meter.columns = column_names
Eng_Cntr_Educ_Blk_meter['A'] = pd.to_datetime(Eng_Cntr_Educ_Blk_meter['A'], unit='D', origin='1899-12-30')

Eng_Cntr_Educ_Blk_meter.rename(columns={"A": "Eng_Cntr_Educ_Blk"}, inplace=True)
Eng_Cntr_Educ_Blk_meter['Total_Day'] = Eng_Cntr_Educ_Blk_meter['24:00'] - Eng_Cntr_Educ_Blk_meter['00:15']

# now let's write it as xlsx file format
# first we need to set up some format styles to be applied to the Excel sheet date column using the named_style module
ewb = pd.ExcelWriter(base_dest_path + "/" + "Eng_Cntr_Educ_Blk_meter.xlsx", engine="openpyxl")
nsmmyy = NamedStyle(name="cd1", number_format="DD-MM-YY")
nsmmmyy = NamedStyle(name="cd2", number_format="MMM-YY")
nsbyy = NamedStyle(name="cd3", number_format="MMMM-YY")
# now that the style has been set up for Excel Writer we can  write the file.
Eng_Cntr_Educ_Blk_meter.to_excel(excel_writer=ewb, sheet_name="Eng_Cntr_Educ_Blk_mtr")
# after this we have to format the Eng_Cntr_Educ_Blk_meter date column using the NamedStyles set up
ws = ewb.book["Eng_Cntr_Educ_Blk_mtr"]
for i in range(1, len(Eng_Cntr_Educ_Blk_meter) + 2):
    ws.cell(row=i, column=2).style = nsmmyy
    # ws.cell(row=i, column=3).style = nsmmmyy
    # ws.cell(row=i, column=2).style = nsbyy
ewb.close()

# Finally we need to sort out the Eng_Cntr_Educ_Blk_meter column width value in the completed Excel sheet
worksheet = openpyxl.load_workbook(base_dest_path + "/" + "Eng_Cntr_Educ_Blk_meter.xlsx")
sheet = worksheet.active
sheet.column_dimensions['B'].width = 19
worksheet.save(base_dest_path + "/" + "Eng_Cntr_Educ_Blk_meter.xlsx")
worksheet.close()

# next meter

Eng_Cntr_Unit_9_meter = pd.read_csv(base_path + "/" + file_name49, index_col=False, skiprows=[0])
Eng_Cntr_Unit_9_meter.columns = column_names
Eng_Cntr_Unit_9_meter['A'] = pd.to_datetime(Eng_Cntr_Unit_9_meter['A'], unit='D', origin='1899-12-30')

Eng_Cntr_Unit_9_meter.rename(columns={"A": "Eng_Cntr_Unit_9"}, inplace=True)
Eng_Cntr_Unit_9_meter['Total_Day'] = Eng_Cntr_Unit_9_meter['24:00'] - Eng_Cntr_Unit_9_meter['00:15']

# now let's write it as xlsx file format
# first we need to set up some format styles to be applied to the Excel sheet date column using the named_style module
ewb = pd.ExcelWriter(base_dest_path + "/" + "Eng_Cntr_Unit_9_meter.xlsx", engine="openpyxl")
nsmmyy = NamedStyle(name="cd1", number_format="DD-MM-YY")
nsmmmyy = NamedStyle(name="cd2", number_format="MMM-YY")
nsbyy = NamedStyle(name="cd3", number_format="MMMM-YY")
# now that the style has been set up for Excel Writer we can  write the file.
Eng_Cntr_Unit_9_meter.to_excel(excel_writer=ewb, sheet_name="Eng_Cntr_Unit_9_mtr")
# after this we have to format the Eng_Cntr_Unit_9_meter date column using the NamedStyles set up
ws = ewb.book["Eng_Cntr_Unit_9_mtr"]
for i in range(1, len(Eng_Cntr_Unit_9_meter) + 2):
    ws.cell(row=i, column=2).style = nsmmyy
    # ws.cell(row=i, column=3).style = nsmmmyy
    # ws.cell(row=i, column=2).style = nsbyy
ewb.close()

# Finally we need to sort out the Eng_Cntr_Unit_9_meter column width value in the completed Excel sheet
worksheet = openpyxl.load_workbook(base_dest_path + "/" + "Eng_Cntr_Unit_9_meter.xlsx")
sheet = worksheet.active
sheet.column_dimensions['B'].width = 19
worksheet.save(base_dest_path + "/" + "Eng_Cntr_Unit_9_meter.xlsx")
worksheet.close()

# next meter

Eng_Cntr_Trinity_Hse_meter = pd.read_csv(base_path + "/" + file_name50, index_col=False, skiprows=[0])
Eng_Cntr_Trinity_Hse_meter.columns = column_names
Eng_Cntr_Trinity_Hse_meter['A'] = pd.to_datetime(Eng_Cntr_Trinity_Hse_meter['A'], unit='D', origin='1899-12-30')

Eng_Cntr_Trinity_Hse_meter.rename(columns={"A": "Eng_Cntr_Trinity_Hse"}, inplace=True)
Eng_Cntr_Trinity_Hse_meter['Total_Day'] = Eng_Cntr_Trinity_Hse_meter['24:00'] - Eng_Cntr_Trinity_Hse_meter['00:15']

# now let's write it as xlsx file format
# first we need to set up some format styles to be applied to the Excel sheet date column using the named_style module
ewb = pd.ExcelWriter(base_dest_path + "/" + "Eng_Cntr_Trinity_Hse_meter.xlsx", engine="openpyxl")
nsmmyy = NamedStyle(name="cd1", number_format="DD-MM-YY")
nsmmmyy = NamedStyle(name="cd2", number_format="MMM-YY")
nsbyy = NamedStyle(name="cd3", number_format="MMMM-YY")
# now that the style has been set up for Excel Writer we can  write the file.
Eng_Cntr_Trinity_Hse_meter.to_excel(excel_writer=ewb, sheet_name="Eng_Cntr_Trin_Hse_mtr")
# after this we have to format the Eng_Cntr_Trinity_Hse_meter date column using the NamedStyles set up
ws = ewb.book["Eng_Cntr_Trin_Hse_mtr"]
for i in range(1, len(Eng_Cntr_Trinity_Hse_meter) + 2):
    ws.cell(row=i, column=2).style = nsmmyy
    # ws.cell(row=i, column=3).style = nsmmmyy
    # ws.cell(row=i, column=2).style = nsbyy
ewb.close()

# Finally we need to sort out the Eng_Cntr_Trinity_Hse_meter column width value in the completed Excel sheet
worksheet = openpyxl.load_workbook(base_dest_path + "/" + "Eng_Cntr_Trinity_Hse_meter.xlsx")
sheet = worksheet.active
sheet.column_dimensions['B'].width = 19
worksheet.save(base_dest_path + "/" + "Eng_Cntr_Trinity_Hse_meter.xlsx")
worksheet.close()

# next meter

Eng_Cntr_Unit_6_meter = pd.read_csv(base_path + "/" + file_name51, index_col=False, skiprows=[0])
Eng_Cntr_Unit_6_meter.columns = column_names
Eng_Cntr_Unit_6_meter['A'] = pd.to_datetime(Eng_Cntr_Unit_6_meter['A'], unit='D', origin='1899-12-30')

Eng_Cntr_Unit_6_meter.rename(columns={"A": "Eng_Cntr_Unit_6"}, inplace=True)
Eng_Cntr_Unit_6_meter['Total_Day'] = Eng_Cntr_Unit_6_meter['24:00'] - Eng_Cntr_Unit_6_meter['00:15']

# now let's write it as xlsx file format
# first we need to set up some format styles to be applied to the Excel sheet date column using the named_style module
ewb = pd.ExcelWriter(base_dest_path + "/" + "Eng_Cntr_Unit_6_meter.xlsx", engine="openpyxl")
nsmmyy = NamedStyle(name="cd1", number_format="DD-MM-YY")
nsmmmyy = NamedStyle(name="cd2", number_format="MMM-YY")
nsbyy = NamedStyle(name="cd3", number_format="MMMM-YY")
# now that the style has been set up for Excel Writer we can  write the file.
Eng_Cntr_Unit_6_meter.to_excel(excel_writer=ewb, sheet_name="Eng_Cntr_Unit_6_mtr")
# after this we have to format the Eng_Cntr_Unit_6_meter date column using the NamedStyles set up
ws = ewb.book["Eng_Cntr_Unit_6_mtr"]
for i in range(1, len(Eng_Cntr_Unit_6_meter) + 2):
    ws.cell(row=i, column=2).style = nsmmyy
    # ws.cell(row=i, column=3).style = nsmmmyy
    # ws.cell(row=i, column=2).style = nsbyy
ewb.close()

# Finally we need to sort out the Eng_Cntr_Unit_6_meter column width value in the completed Excel sheet
worksheet = openpyxl.load_workbook(base_dest_path + "/" + "Eng_Cntr_Unit_6_meter.xlsx")
sheet = worksheet.active
sheet.column_dimensions['B'].width = 19
worksheet.save(base_dest_path + "/" + "Eng_Cntr_Unit_6_meter.xlsx")
worksheet.close()

# next meter

Trinity_Hse_Light_meter = pd.read_csv(base_path + "/" + file_name52, index_col=False, skiprows=[0])
Trinity_Hse_Light_meter.columns = column_names
Trinity_Hse_Light_meter['A'] = pd.to_datetime(Trinity_Hse_Light_meter['A'], unit='D', origin='1899-12-30')

Trinity_Hse_Light_meter.rename(columns={"A": "Trinity_Hse_Light"}, inplace=True)
Trinity_Hse_Light_meter['Total_Day'] = Trinity_Hse_Light_meter['24:00'] - Trinity_Hse_Light_meter['00:15']

# now let's write it as xlsx file format
# first we need to set up some format styles to be applied to the Excel sheet date column using the named_style module
ewb = pd.ExcelWriter(base_dest_path + "/" + "Trinity_Hse_Light_meter.xlsx", engine="openpyxl")
nsmmyy = NamedStyle(name="cd1", number_format="DD-MM-YY")
nsmmmyy = NamedStyle(name="cd2", number_format="MMM-YY")
nsbyy = NamedStyle(name="cd3", number_format="MMMM-YY")
# now that the style has been set up for Excel Writer we can  write the file.
Trinity_Hse_Light_meter.to_excel(excel_writer=ewb, sheet_name="Trinity_Hse_Light_mtr")
# after this we have to format the Trinity_Hse_Light_meter date column using the NamedStyles set up
ws = ewb.book["Trinity_Hse_Light_mtr"]
for i in range(1, len(Trinity_Hse_Light_meter) + 2):
    ws.cell(row=i, column=2).style = nsmmyy
    # ws.cell(row=i, column=3).style = nsmmmyy
    # ws.cell(row=i, column=2).style = nsbyy
ewb.close()

# Finally we need to sort out the Trinity_Hse_Light_meter column width value in the completed Excel sheet
worksheet = openpyxl.load_workbook(base_dest_path + "/" + "Trinity_Hse_Light_meter.xlsx")
sheet = worksheet.active
sheet.column_dimensions['B'].width = 19
worksheet.save(base_dest_path + "/" + "Trinity_Hse_Light_meter.xlsx")
worksheet.close()

# next meter

Trinity_Hse_GS_meter = pd.read_csv(base_path + "/" + file_name53, index_col=False, skiprows=[0])
Trinity_Hse_GS_meter.columns = column_names
Trinity_Hse_GS_meter['A'] = pd.to_datetime(Trinity_Hse_GS_meter['A'], unit='D', origin='1899-12-30')

Trinity_Hse_GS_meter.rename(columns={"A": "Trinity_Hse_GS"}, inplace=True)
Trinity_Hse_GS_meter['Total_Day'] = Trinity_Hse_GS_meter['24:00'] - Trinity_Hse_GS_meter['00:15']

# now let's write it as xlsx file format
# first we need to set up some format styles to be applied to the Excel sheet date column using the named_style module
ewb = pd.ExcelWriter(base_dest_path + "/" + "Trinity_Hse_GS_meter.xlsx", engine="openpyxl")
nsmmyy = NamedStyle(name="cd1", number_format="DD-MM-YY")
nsmmmyy = NamedStyle(name="cd2", number_format="MMM-YY")
nsbyy = NamedStyle(name="cd3", number_format="MMMM-YY")
# now that the style has been set up for Excel Writer we can  write the file.
Trinity_Hse_GS_meter.to_excel(excel_writer=ewb, sheet_name="Trinity_Hse_GS_mtr")
# after this we have to format the Trinity_Hse_GS_meter date column using the NamedStyles set up
ws = ewb.book["Trinity_Hse_GS_mtr"]
for i in range(1, len(Trinity_Hse_GS_meter) + 2):
    ws.cell(row=i, column=2).style = nsmmyy
    # ws.cell(row=i, column=3).style = nsmmmyy
    # ws.cell(row=i, column=2).style = nsbyy
ewb.close()

# Finally we need to sort out the Trinity_Hse_GS_meter column width value in the completed Excel sheet
worksheet = openpyxl.load_workbook(base_dest_path + "/" + "Trinity_Hse_GS_meter.xlsx")
sheet = worksheet.active
sheet.column_dimensions['B'].width = 19
worksheet.save(base_dest_path + "/" + "Trinity_Hse_GS_meter.xlsx")
worksheet.close()

# next meter

Trinity_Hse_Water_meter = pd.read_csv(base_path + "/" + file_name54, index_col=False, skiprows=[0])
Trinity_Hse_Water_meter.columns = column_names
Trinity_Hse_Water_meter['A'] = pd.to_datetime(Trinity_Hse_Water_meter['A'], unit='D', origin='1899-12-30')

Trinity_Hse_Water_meter.rename(columns={"A": "Trinity_Hse_Water"}, inplace=True)
Trinity_Hse_Water_meter['Total_Day'] = Trinity_Hse_Water_meter['24:00'] - Trinity_Hse_Water_meter['00:15']

# now let's write it as xlsx file format
# first we need to set up some format styles to be applied to the Excel sheet date column using the named_style module
ewb = pd.ExcelWriter(base_dest_path + "/" + "Trinity_Hse_Water_meter.xlsx", engine="openpyxl")
nsmmyy = NamedStyle(name="cd1", number_format="DD-MM-YY")
nsmmmyy = NamedStyle(name="cd2", number_format="MMM-YY")
nsbyy = NamedStyle(name="cd3", number_format="MMMM-YY")
# now that the style has been set up for Excel Writer we can  write the file.
Trinity_Hse_Water_meter.to_excel(excel_writer=ewb, sheet_name="Trinity_Hse_Wtr_mtr")
# after this we have to format the Trinity_Hse_Water_meter date column using the NamedStyles set up
ws = ewb.book["Trinity_Hse_Wtr_mtr"]
for i in range(1, len(Trinity_Hse_Water_meter) + 2):
    ws.cell(row=i, column=2).style = nsmmyy
    # ws.cell(row=i, column=3).style = nsmmmyy
    # ws.cell(row=i, column=2).style = nsbyy
ewb.close()

# Finally we need to sort out the Trinity_Hse_Water_meter column width value in the completed Excel sheet
worksheet = openpyxl.load_workbook(base_dest_path + "/" + "Trinity_Hse_Water_meter.xlsx")
sheet = worksheet.active
sheet.column_dimensions['B'].width = 19
worksheet.save(base_dest_path + "/" + "Trinity_Hse_Water_meter.xlsx")
worksheet.close()
# now run the process completed message
show_message()
